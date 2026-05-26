"""Manejo del Excel de entrada y salida para el proceso QA de servicios."""
import os
from datetime import datetime
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

EXCEL_MAX_CHARS = 32767

# Mapeo de columnas (nombre → letra de columna en Excel)
COL_MAP = {
    "Ejecuta": "A",
    "IdMetodo": "B",
    "NombreRPC2": "C",
    "NombreStorm": "D",
    "NombreGPL": "E",
    "Comentario": "S",
    "LlamadaGPLyStorm": "T",
    "LlamadaRPC2": "U",
    "FmtLlamada": "V",
    "FmtRespuesta": "W",
    "CodRespRPC2": "X",
    "RespRPC2": "Y",
    "CodRespGPL": "Z",
    "RespGPL": "AA",
    "CodRespStorm": "AB",
    "RespStorm": "AC",
    "HoraEjecucion": "AD",
    "ContRespRPC2": "AE",
    "ContRespGPL": "AF",
    "ContRespSTORM": "AG",
    "GPLIncluyeColRPC2": "AH",
    "GPLIgualRPC2": "AI",
    "GPLIncluyeColSTORM": "AJ",
    "GPLIgualSTORM": "AK",
}


def leer_excel_entrada(ruta_archivo: str):
    """Abre el libro Excel de entrada y retorna el workbook (data_only para valores calculados)."""
    wb = load_workbook(ruta_archivo, data_only=True)
    return wb


def crear_excel_salida(wb_entrada, ruta_salida: str):
    """Crea un nuevo Excel copiando los datos de la hoja activa del wb de entrada.
    
    Copia valores (no fórmulas) para que sea independiente del original.
    """
    ws_entrada = wb_entrada.active
    wb_salida = Workbook()
    ws_salida = wb_salida.active
    ws_salida.title = ws_entrada.title

    for row in ws_entrada.iter_rows(min_row=1, max_row=ws_entrada.max_row,
                                     max_col=ws_entrada.max_column):
        for cell in row:
            new_cell = ws_salida.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copiar anchos de columna
    for col_letter, dim in ws_entrada.column_dimensions.items():
        ws_salida.column_dimensions[col_letter].width = dim.width

    os.makedirs(os.path.dirname(ruta_salida) or ".", exist_ok=True)
    wb_salida.save(ruta_salida)
    return wb_salida


def nombre_archivo_salida(ruta_directorio: str) -> str:
    """Genera el nombre del archivo de salida con fecha actual."""
    fecha = datetime.now().strftime("%Y%m%d")
    nombre = f"ListaMetodosServicioQAEjecutados{fecha}.xlsx"
    return os.path.join(ruta_directorio, nombre)


def leer_fila(ws, num_fila: int) -> dict:
    """Lee una fila del Excel y retorna un diccionario con los valores relevantes."""
    def val(col_letter):
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)
        return ws.cell(row=num_fila, column=col_idx).value

    return {
        "Ejecuta": val(COL_MAP["Ejecuta"]),
        "IdMetodo": val(COL_MAP["IdMetodo"]),
        "NombreRPC2": val(COL_MAP["NombreRPC2"]),
        "NombreStorm": val(COL_MAP["NombreStorm"]),
        "NombreGPL": val(COL_MAP["NombreGPL"]),
        "Comentario": val(COL_MAP["Comentario"]) or "",
        "LlamadaGPLyStorm": val(COL_MAP["LlamadaGPLyStorm"]),
        "LlamadaRPC2": val(COL_MAP["LlamadaRPC2"]),
        "FmtLlamada": val(COL_MAP["FmtLlamada"]) or "application/json",
        "FmtRespuesta": val(COL_MAP["FmtRespuesta"]) or "application/json",
    }


def escribir_resultado(ws, num_fila: int, campo: str, valor):
    """Escribe un valor en la celda correspondiente al campo indicado."""
    from openpyxl.utils import column_index_from_string
    col_letter = COL_MAP.get(campo)
    if not col_letter:
        return
    col_idx = column_index_from_string(col_letter)
    ws.cell(row=num_fila, column=col_idx, value=valor)


def escribir_respuesta_con_desborde(ws, num_fila: int, campo_resp: str,
                                     campo_cont: str, texto: str):
    """Escribe la respuesta en la columna principal y desborda si excede el límite de Excel.
    
    Si el texto excede EXCEL_MAX_CHARS, se divide entre la columna principal
    y la columna de continuación. Si aún desborda, se corta con indicador.
    """
    if texto is None:
        texto = ""
    texto = str(texto)

    if len(texto) <= EXCEL_MAX_CHARS:
        escribir_resultado(ws, num_fila, campo_resp, texto)
        return

    # Primera parte en columna principal
    escribir_resultado(ws, num_fila, campo_resp, texto[:EXCEL_MAX_CHARS])
    resto = texto[EXCEL_MAX_CHARS:]

    # Segunda parte en columna de continuación
    if len(resto) <= EXCEL_MAX_CHARS:
        escribir_resultado(ws, num_fila, campo_cont, resto)
    else:
        # Cortar con indicador
        espacio = EXCEL_MAX_CHARS - 11
        escribir_resultado(ws, num_fila, campo_cont, resto[:espacio] + " cortado...")


def sanitizar_nombre_archivo(nombre: str) -> str:
    """Elimina caracteres no válidos para nombres de archivo."""
    invalidos = '<>:"/\\|?*'
    for c in invalidos:
        nombre = nombre.replace(c, "_")
    return nombre.strip()[:200]
