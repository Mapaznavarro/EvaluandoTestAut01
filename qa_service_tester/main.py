"""
QA Service Tester - Programa principal
=====================================
Reemplazo del flujo de Power Automate para comparación de servicios RPC2, STORM y GPL.

Uso:
    python main.py [ruta_config.ini]

El programa:
1. Lee el Excel con la lista de métodos y parámetros
2. Crea un Excel de salida con fecha
3. Para cada fila con Ejecuta=S, llama a los 3 servicios (RPC2, STORM, GPL)
4. Registra respuestas, compara columnas y contenido
5. Genera archivos de respuesta individuales
6. Graba el Excel de resultados
"""
import os
import sys
import logging
from datetime import datetime

from config import cargar_config
from excel_handler import (
    leer_excel_entrada, crear_excel_salida, nombre_archivo_salida,
    leer_fila, escribir_resultado, escribir_respuesta_con_desborde,
    sanitizar_nombre_archivo
)
from service_caller import llamar_rpc2, llamar_storm, llamar_gpl, RespuestaServicio
from response_parser import parsear_rpc2, parsear_json_gpl, parsear_json_storm
from comparator import columnas_incluidas, comparar_tablas

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger(__name__)


def guardar_archivo_respuesta(ruta_resultados: str, nombre_metodo: str,
                               escenario: str, contenido: str, extension: str = ""):
    """Guarda la respuesta de un servicio en un archivo individual."""
    nombre = sanitizar_nombre_archivo(f"{nombre_metodo}_{escenario}")
    if extension:
        nombre = f"{nombre}.{extension}"
    ruta = os.path.join(ruta_resultados, nombre)
    os.makedirs(ruta_resultados, exist_ok=True)
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(contenido)
    return ruta


def procesar_fila(ws, num_fila: int, datos_fila: dict, cfg: dict, ruta_resultados: str):
    """Procesa una fila del Excel: llama a los 3 servicios, registra y compara."""
    nombre_rpc2 = datos_fila["NombreRPC2"] or ""
    nombre_storm = datos_fila["NombreStorm"] or ""
    nombre_gpl = datos_fila["NombreGPL"] or ""
    comentario = datos_fila["Comentario"] or ""
    body_gpl_storm = datos_fila["LlamadaGPLyStorm"] or ""
    body_rpc2 = datos_fila["LlamadaRPC2"] or ""
    fmt_llamada = datos_fila["FmtLlamada"]
    fmt_respuesta = datos_fila["FmtRespuesta"]

    log.info(f"  Fila {num_fila}: RPC2={nombre_rpc2}, STORM={nombre_storm}, "
             f"GPL={nombre_gpl} [{comentario}]")

    resp_rpc2 = RespuestaServicio(disponible=False, error="Sin nombre de método RPC2")
    resp_storm = RespuestaServicio(disponible=False, error="Sin nombre de método STORM")
    resp_gpl = RespuestaServicio(disponible=False, error="Sin nombre de método GPL")
    datos_rpc2 = None
    datos_storm = None
    datos_gpl = None

    # ── Paso 4.1: Llamar RPC2 (SOAP) ──
    if nombre_rpc2 and body_rpc2:
        log.info(f"    Llamando RPC2: {nombre_rpc2}...")
        resp_rpc2 = llamar_rpc2(
            url=cfg["URL_Servicio_RPC2"],
            body_xml=body_rpc2,
            timeout=cfg["TIEMPO_ESPERA_RPC"]
        )
        if resp_rpc2.disponible and not resp_rpc2.error:
            datos_rpc2 = parsear_rpc2(resp_rpc2.cuerpo_respuesta)
            log.info(f"    RPC2 código={resp_rpc2.codigo_respuesta}, "
                     f"cols={len(datos_rpc2.columnas)}, filas={len(datos_rpc2.filas)}, "
                     f"error={datos_rpc2.es_error}")
        else:
            log.warning(f"    RPC2 error: {resp_rpc2.error}")
    else:
        log.info(f"    RPC2 omitido (sin datos)")

    escribir_resultado(ws, num_fila, "CodRespRPC2",
                       resp_rpc2.codigo_respuesta if resp_rpc2.disponible else resp_rpc2.error)
    texto_rpc2 = resp_rpc2.cuerpo_respuesta if resp_rpc2.disponible and not resp_rpc2.error else resp_rpc2.error
    escribir_respuesta_con_desborde(ws, num_fila, "RespRPC2", "ContRespRPC2", texto_rpc2)

    if nombre_rpc2 and resp_rpc2.cuerpo_respuesta:
        ext = "xml"
        guardar_archivo_respuesta(ruta_resultados, nombre_rpc2, comentario,
                                   resp_rpc2.cuerpo_respuesta, ext)

    # ── Paso 4.2: Llamar STORM (REST) ──
    if nombre_storm and body_gpl_storm:
        log.info(f"    Llamando STORM: {nombre_storm}...")
        resp_storm = llamar_storm(
            url_base=cfg["URL_Servicio_STORM"],
            nombre_metodo=nombre_storm,
            body=body_gpl_storm,
            timeout=cfg["TIEMPO_ESPERA_STORM"],
            content_type=fmt_llamada,
            accept=fmt_respuesta,
            usuario=cfg["USUARIO_STORM"],
            password=cfg["PASSWORD_STORM"]
        )
        if resp_storm.disponible and not resp_storm.error:
            datos_storm = parsear_json_storm(resp_storm.cuerpo_respuesta)
            log.info(f"    STORM código={resp_storm.codigo_respuesta}, "
                     f"cols={len(datos_storm.columnas)}, filas={len(datos_storm.filas)}, "
                     f"error={datos_storm.es_error}")
        else:
            log.warning(f"    STORM error: {resp_storm.error}")
    else:
        log.info(f"    STORM omitido (sin datos)")

    escribir_resultado(ws, num_fila, "CodRespStorm",
                       resp_storm.codigo_respuesta if resp_storm.disponible else resp_storm.error)
    texto_storm = resp_storm.cuerpo_respuesta if resp_storm.disponible and not resp_storm.error else resp_storm.error
    escribir_respuesta_con_desborde(ws, num_fila, "RespStorm", "ContRespSTORM", texto_storm)

    if nombre_storm and resp_storm.cuerpo_respuesta:
        ext = "json" if "json" in fmt_respuesta.lower() else "xml"
        guardar_archivo_respuesta(ruta_resultados, nombre_storm, comentario,
                                   resp_storm.cuerpo_respuesta, ext)

    # ── Paso 4.3: Llamar GPL (REST) ──
    if nombre_gpl and body_gpl_storm:
        log.info(f"    Llamando GPL: {nombre_gpl}...")
        resp_gpl = llamar_gpl(
            url_base=cfg["URL_Servicio_GPL"],
            nombre_metodo=nombre_gpl,
            body=body_gpl_storm,
            timeout=cfg["TIEMPO_ESPERA_GPL"],
            content_type=fmt_llamada,
            accept=fmt_respuesta,
            usuario=cfg["USUARIO_GPL"],
            password=cfg["PASSWORD_GPL"]
        )
        if resp_gpl.disponible and not resp_gpl.error:
            datos_gpl = parsear_json_gpl(resp_gpl.cuerpo_respuesta)
            log.info(f"    GPL código={resp_gpl.codigo_respuesta}, "
                     f"cols={len(datos_gpl.columnas)}, filas={len(datos_gpl.filas)}, "
                     f"error={datos_gpl.es_error}")
        else:
            log.warning(f"    GPL error: {resp_gpl.error}")
    else:
        log.info(f"    GPL omitido (sin datos)")

    escribir_resultado(ws, num_fila, "CodRespGPL",
                       resp_gpl.codigo_respuesta if resp_gpl.disponible else resp_gpl.error)
    texto_gpl = resp_gpl.cuerpo_respuesta if resp_gpl.disponible and not resp_gpl.error else resp_gpl.error
    escribir_respuesta_con_desborde(ws, num_fila, "RespGPL", "ContRespGPL", texto_gpl)

    if nombre_gpl and resp_gpl.cuerpo_respuesta:
        ext = "json" if "json" in fmt_respuesta.lower() else "xml"
        guardar_archivo_respuesta(ruta_resultados, nombre_gpl.replace("/", "_"),
                                   comentario, resp_gpl.cuerpo_respuesta, ext)

    # ── Paso 4.4: Comparar columnas GPL vs RPC2 ──
    if (datos_gpl and datos_rpc2 and datos_gpl.tiene_datos and datos_rpc2.tiene_datos):
        incluye = columnas_incluidas(datos_rpc2, datos_gpl)
        escribir_resultado(ws, num_fila, "GPLIncluyeColRPC2", "SI" if incluye else "NO")
    elif not resp_rpc2.disponible:
        escribir_resultado(ws, num_fila, "GPLIncluyeColRPC2", "RPC2 no disponible")
    elif not resp_gpl.disponible:
        escribir_resultado(ws, num_fila, "GPLIncluyeColRPC2", "GPL no disponible")
    elif datos_rpc2 and datos_rpc2.es_error:
        escribir_resultado(ws, num_fila, "GPLIncluyeColRPC2", f"Error RPC2: {datos_rpc2.mensaje_error[:80]}")
    elif datos_gpl and datos_gpl.es_error:
        escribir_resultado(ws, num_fila, "GPLIncluyeColRPC2", f"Error GPL: {datos_gpl.mensaje_error[:80]}")
    else:
        escribir_resultado(ws, num_fila, "GPLIncluyeColRPC2", "N/A")

    # ── Paso 4.5: Comparar contenido GPL vs RPC2 ──
    from openpyxl.utils import column_index_from_string
    gpl_incluye_rpc2 = ws.cell(row=num_fila,
                                column=column_index_from_string("AH")).value
    if gpl_incluye_rpc2 == "SI" and datos_gpl and datos_rpc2:
        nombre_base_rpc = sanitizar_nombre_archivo(f"{nombre_rpc2}_{comentario}")
        nombre_base_gpl = sanitizar_nombre_archivo(
            f"{nombre_gpl.replace('/', '_')}_{comentario}")
        csv_rpc = os.path.join(ruta_resultados, f"{nombre_base_rpc}.csv")
        csv_gpl = os.path.join(ruta_resultados, f"{nombre_base_gpl}.csv")
        resultado_cmp = comparar_tablas(datos_rpc2, datos_gpl, csv_rpc, csv_gpl)
        escribir_resultado(ws, num_fila, "GPLIgualRPC2", resultado_cmp)
    elif gpl_incluye_rpc2 and gpl_incluye_rpc2 != "SI":
        escribir_resultado(ws, num_fila, "GPLIgualRPC2", "N/A (cols no incluidas)")

    # ── Comparar columnas GPL vs STORM ──
    if (datos_gpl and datos_storm and datos_gpl.tiene_datos and datos_storm.tiene_datos):
        incluye_storm = columnas_incluidas(datos_storm, datos_gpl)
        escribir_resultado(ws, num_fila, "GPLIncluyeColSTORM", "SI" if incluye_storm else "NO")

        if incluye_storm:
            nombre_base_storm = sanitizar_nombre_archivo(f"{nombre_storm}_{comentario}")
            nombre_base_gpl2 = sanitizar_nombre_archivo(
                f"{nombre_gpl.replace('/', '_')}_{comentario}")
            csv_storm = os.path.join(ruta_resultados, f"{nombre_base_storm}_storm.csv")
            csv_gpl2 = os.path.join(ruta_resultados, f"{nombre_base_gpl2}_gpl_vs_storm.csv")
            resultado_cmp_storm = comparar_tablas(datos_storm, datos_gpl, csv_storm, csv_gpl2)
            escribir_resultado(ws, num_fila, "GPLIgualSTORM", resultado_cmp_storm)
        else:
            escribir_resultado(ws, num_fila, "GPLIgualSTORM", "N/A (cols no incluidas)")
    elif not resp_storm.disponible:
        escribir_resultado(ws, num_fila, "GPLIncluyeColSTORM", "STORM no disponible")
    elif not resp_gpl.disponible:
        escribir_resultado(ws, num_fila, "GPLIncluyeColSTORM", "GPL no disponible")
    elif datos_storm and datos_storm.es_error:
        escribir_resultado(ws, num_fila, "GPLIncluyeColSTORM",
                           f"Error STORM: {datos_storm.mensaje_error[:80]}")
    elif datos_gpl and datos_gpl.es_error:
        escribir_resultado(ws, num_fila, "GPLIncluyeColSTORM",
                           f"Error GPL: {datos_gpl.mensaje_error[:80]}")
    else:
        escribir_resultado(ws, num_fila, "GPLIncluyeColSTORM", "N/A")

    # ── Paso 4.6: Hora de ejecución ──
    escribir_resultado(ws, num_fila, "HoraEjecucion",
                       datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


def main(ruta_config: str = None):
    """Punto de entrada principal del programa."""
    log.info("=" * 70)
    log.info("QA Service Tester - Comparador de servicios RPC2, STORM y GPL")
    log.info("=" * 70)

    # Cargar configuración
    cfg = cargar_config(ruta_config)
    log.info(f"Configuración cargada:")
    log.info(f"  RPC2:  {cfg['URL_Servicio_RPC2']} (timeout={cfg['TIEMPO_ESPERA_RPC']}s)")
    log.info(f"  STORM: {cfg['URL_Servicio_STORM']} (timeout={cfg['TIEMPO_ESPERA_STORM']}s)")
    log.info(f"  GPL:   {cfg['URL_Servicio_GPL']} (timeout={cfg['TIEMPO_ESPERA_GPL']}s)")

    ruta_excel = cfg["ARCHIVO_EXCEL_ENTRADA"]
    ruta_resultados = cfg["RUTA_RESULTADOS"]
    os.makedirs(ruta_resultados, exist_ok=True)

    # Paso 1: Abrir Excel de entrada
    log.info(f"\nPaso 1: Abriendo Excel de entrada: {ruta_excel}")
    if not os.path.exists(ruta_excel):
        log.error(f"Archivo no encontrado: {ruta_excel}")
        sys.exit(1)
    wb_entrada = leer_excel_entrada(ruta_excel)
    ws_entrada = wb_entrada.active
    log.info(f"  Hoja activa: '{ws_entrada.title}', "
             f"Filas: {ws_entrada.max_row}, Columnas: {ws_entrada.max_column}")

    # Paso 2: Crear Excel de salida
    ruta_salida = nombre_archivo_salida(ruta_resultados)
    log.info(f"\nPaso 2: Creando Excel de salida: {ruta_salida}")
    wb_salida = crear_excel_salida(wb_entrada, ruta_salida)
    ws_salida = wb_salida.active

    # Paso 3 y 4: Recorrer filas
    total_filas = ws_salida.max_row
    filas_ejecutadas = 0
    filas_omitidas = 0

    log.info(f"\nPaso 3-4: Procesando {total_filas - 1} filas de datos...")
    for num_fila in range(2, total_filas + 1):
        datos = leer_fila(ws_salida, num_fila)

        # Paso 4.0: Verificar Ejecuta
        ejecuta = str(datos["Ejecuta"] or "").strip().upper()
        if ejecuta == "FIN":
            log.info(f"  Fila {num_fila}: encontrado 'fin', deteniendo recorrido.")
            break
        if ejecuta != "S":
            filas_omitidas += 1
            continue

        filas_ejecutadas += 1
        procesar_fila(ws_salida, num_fila, datos, cfg, ruta_resultados)

        # Guardar periódicamente (cada 20 filas)
        if filas_ejecutadas % 20 == 0:
            log.info(f"  Guardando progreso parcial ({filas_ejecutadas} filas procesadas)...")
            wb_salida.save(ruta_salida)

    # Paso 4.8: Guardar Excel final
    log.info(f"\nPaso 4.8: Guardando Excel final...")
    wb_salida.save(ruta_salida)

    log.info(f"\n{'=' * 70}")
    log.info(f"RESUMEN:")
    log.info(f"  Filas ejecutadas: {filas_ejecutadas}")
    log.info(f"  Filas omitidas:   {filas_omitidas}")
    log.info(f"  Excel de salida:  {ruta_salida}")
    log.info(f"  Archivos en:      {ruta_resultados}")
    log.info(f"{'=' * 70}")


if __name__ == "__main__":
    ruta_cfg = sys.argv[1] if len(sys.argv) > 1 else None
    main(ruta_cfg)
