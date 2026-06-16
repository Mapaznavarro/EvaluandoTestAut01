"""
body_builder.py
===============
Arma el body JSON de llamadas a GPL/STORM con varios niveles (objetos anidados
y/o arreglos), leyendo del Excel:

  * la ESTRUCTURA (ruta + tipo) una sola vez por método, en la hoja BaseParametros;
  * los VALORES horizontales, una fila por llamada, en la hoja de casos.

Es ADITIVO: solo se usa para filas marcadas con Nivelado = "SI". Las llamadas
planas que ya funcionan (columna "LlamadaGPL y Storm" armada por paramGPL) no se
tocan. No modifica paramGPL ni paramRPC2.

Notación de ruta (columna "Atributo" de BaseParametros para métodos nivelados):
    instrumentos.codigo          -> {"instrumentos": {"codigo": ...}}
    operaciones[0].importe       -> {"operaciones": [ {"importe": ...} ]}
El "Tipo" decide la coacción y json.dumps pone comillas y escapes.
"""
import json
import re
from openpyxl.utils import column_index_from_string as _idx

# ============================================================================
# AJUSTAR según tu Excel (ya corregido)
# ----------------------------------------------------------------------------
# ESTRUCTURA del método (ruta + tipo): se define UNA VEZ por método.
HOJA_CATALOGO   = "BaseParametros"  # hoja con la estructura
CAT_COL_ID      = "A"               # Id Metodo (solo en la 1ª fila del bloque)
CAT_COL_RUTA    = "C"               # Atributo / Ruta (notación con puntos)
CAT_COL_TIPO    = "D"               # Tipo (int/double/boolean/string/dateTime.iso8601)
CAT_FILA_INICIO = 2                 # primera fila de datos del catálogo

# VALORES de cada llamada: para una fila de "Casos Todos" con Niveles='S', los
# valores se buscan en la hoja "Casos Mas Parametros" emparejando Id Metodo +
# Comentario, y se leen en horizontal desde la columna J.
HOJA_VALORES       = "Casos Mas Parametros"
VAL_COL_ID         = "A"   # Id Metodo en "Casos Mas Parametros"
VAL_COL_COMENTARIO = "C"   # Comentario (escenario) en "Casos Mas Parametros"
VAL_COL_INICIO     = "J"   # primera columna de valores (Parametro 1)
VAL_FILA_INICIO    = 14    # primera fila de datos en "Casos Mas Parametros"
# ============================================================================

_SEG_ARREGLO = re.compile(r"^(.+?)\[(\d+)\]$")


def es_nivelada(valor):
    """Indica si la fila debe armar un body JSON de varios niveles en Python.

    Acepta el contenido de la columna "Niveles":
      * vacío, "1", "NO", "0"  -> False (body plano actual, columna LlamadaGPLyStorm)
      * "SI" / "S" / "TRUE"    -> True
      * un número >= 2         -> True (cantidad de niveles)
    """
    s = str(valor or "").strip().upper()
    if s in ("", "NO", "N", "0", "1", "FALSE"):
        return False
    if s in ("SI", "SÍ", "S", "TRUE", "YES", "Y"):
        return True
    try:
        return float(s.replace(",", ".")) >= 2
    except ValueError:
        return False


def _quita_comillas(s):
    s = str(s).strip()
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1]
    return s


def coaccionar(tipo, valor):
    """Convierte el valor de celda al tipo Python según 'Tipo'.

    Quita primero las comillas envolventes (los valores en "Casos Mas Parametros"
    vienen como "110", "18000033", etc.) y luego tipa.
    """
    t = (tipo or "string").strip().lower()
    if valor is None:
        return None
    s = _quita_comillas(valor)
    if s == "":
        return "" if t in ("string", "datetime.iso8601") else None
    if t == "int":
        return int(float(s))
    if t == "double":
        return float(s.replace(",", "."))
    if t == "boolean":
        return s.strip().lower() in ("1", "true", "s", "si", "sí", "verdadero")
    return s                                # string, dateTime.iso8601


def _asignar(raiz, ruta, valor):
    """Inserta 'valor' en 'raiz' siguiendo la ruta con puntos y/o [indice]."""
    segmentos = _quita_comillas(ruta).split(".")
    nodo = raiz
    for i, seg in enumerate(segmentos):
        seg = seg.strip()
        es_ultimo = (i == len(segmentos) - 1)
        m = _SEG_ARREGLO.match(seg)
        if m:                                   # segmento de arreglo: nombre[idx]
            nombre, k = m.group(1), int(m.group(2))
            lista = nodo.setdefault(nombre, [])
            if not isinstance(lista, list):
                raise ValueError(f"'{nombre}' se usa como arreglo y como objeto en '{ruta}'")
            while len(lista) <= k:
                lista.append({})
            if es_ultimo:
                lista[k] = valor
            else:
                if not isinstance(lista[k], dict):
                    lista[k] = {}
                nodo = lista[k]
        else:                                   # segmento de objeto
            if es_ultimo:
                nodo[seg] = valor
            else:
                nodo = nodo.setdefault(seg, {})
                if not isinstance(nodo, dict):
                    raise ValueError(f"'{seg}' se usa como objeto y como valor en '{ruta}'")
    return raiz


def construir_body(definiciones, indent=None):
    """Arma el JSON a partir de una lista de tuplas (ruta, tipo, valor).

    indent=None -> una sola línea (lo que se envía); indent=2 -> legible.
    Devuelve el body JSON como string (UTF-8, sin escapar tildes).
    """
    raiz = {}
    for ruta, tipo, valor in definiciones:
        if ruta is None or str(ruta).strip() == "" or str(ruta).strip().lower() == "fin":
            continue
        _asignar(raiz, str(ruta), coaccionar(tipo, valor))
    return json.dumps(raiz, ensure_ascii=False, indent=indent)


def leer_catalogo(ws_cat, id_metodo):
    """Lee el bloque (ruta, tipo) del método desde BaseParametros.

    Localiza la fila cuyo CAT_COL_ID == id_metodo y lee hacia abajo hasta 'fin'
    (o hasta que comience otro Id Metodo). Devuelve [(ruta, tipo), ...].
    """
    ci, cr, ct = _idx(CAT_COL_ID), _idx(CAT_COL_RUTA), _idx(CAT_COL_TIPO)
    fila_ini = None
    for r in range(CAT_FILA_INICIO, ws_cat.max_row + 1):
        v = ws_cat.cell(row=r, column=ci).value
        if v not in (None, "") and str(v).strip() == str(id_metodo).strip():
            fila_ini = r
            break
    if fila_ini is None:
        raise ValueError(f"No se encontró el Id Metodo {id_metodo} en la hoja {HOJA_CATALOGO}")

    defs = []
    for r in range(fila_ini, ws_cat.max_row + 1):
        ruta = ws_cat.cell(row=r, column=cr).value
        if r > fila_ini:                        # ¿comenzó otro bloque?
            v_id = ws_cat.cell(row=r, column=ci).value
            if v_id not in (None, "") and str(v_id).strip() != str(id_metodo).strip():
                break
        if ruta is None or str(ruta).strip() == "":
            continue
        if str(ruta).strip().lower() == "fin":
            break
        tipo = ws_cat.cell(row=r, column=ct).value
        defs.append((ruta, tipo))
    if not defs:
        raise ValueError(f"El bloque del método {id_metodo} no tiene atributos")
    return defs


def leer_valores(ws_val, num_fila, n, col_inicio=None):
    """Lee n valores horizontales de una fila, empezando en col_inicio (o VAL_COL_INICIO)."""
    ci = _idx(col_inicio or VAL_COL_INICIO)
    return [ws_val.cell(row=num_fila, column=ci + k).value for k in range(n)]


def buscar_fila_valores(ws_val, id_metodo, comentario):
    """Ubica en "Casos Mas Parametros" la fila que empareja Id Metodo + Comentario.

    Devuelve el número de fila. Lanza ValueError si no hay coincidencia.
    """
    ci, cc = _idx(VAL_COL_ID), _idx(VAL_COL_COMENTARIO)
    obj_id = str(id_metodo).strip()
    obj_com = str(comentario or "").strip()
    for r in range(VAL_FILA_INICIO, ws_val.max_row + 1):
        v_id = ws_val.cell(row=r, column=ci).value
        if v_id in (None, ""):
            continue
        if str(v_id).strip() != obj_id:
            continue
        v_com = str(ws_val.cell(row=r, column=cc).value or "").strip()
        if v_com == obj_com:
            return r
    raise ValueError(
        f"No se encontró en '{HOJA_VALORES}' una fila con Id Metodo={id_metodo} "
        f"y Comentario='{comentario}'")


def construir_body_para_caso(ws_cat, ws_val, id_metodo, comentario, indent=None):
    """Punto de entrada principal.

    ws_cat    : hoja BaseParametros (estructura ruta+tipo).
    ws_val    : hoja "Casos Mas Parametros" (valores).
    id_metodo : Id Metodo de la fila de "Casos Todos".
    comentario: Comentario de la fila de "Casos Todos".

    Ambas hojas provienen del libro de ENTRADA, que conserva todas las hojas.
    """
    catalogo = leer_catalogo(ws_cat, id_metodo)
    fila_val = buscar_fila_valores(ws_val, id_metodo, comentario)
    valores = leer_valores(ws_val, fila_val, len(catalogo), VAL_COL_INICIO)
    definiciones = [(ruta, tipo, valores[i]) for i, (ruta, tipo) in enumerate(catalogo)]
    return construir_body(definiciones, indent=indent)
