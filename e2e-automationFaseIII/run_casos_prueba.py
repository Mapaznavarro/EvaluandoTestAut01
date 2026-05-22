"""
run_casos_prueba.py
-------------------
Ejecuta los Casos de Prueba con Datos de Entrada definidos en un Excel.

Lee un Excel con:
  - Camino a la Pantalla  : ruta de menús separados por ' > '
  - Tipo Pantalla         : Consulta / CRUD / Transacciones
  - Acción                : Create / Read / Update / Delete (vacío para Consulta)
  - Dato1..Dato20         : datos formato [Frame].Field.Valor
  - Resultado             : escrito por el programa

Cada caso parte SIEMPRE desde el menú hamburguesa, navega la ruta y
despacha al módulo correspondiente según el tipo de pantalla.

Uso:
    python run_casos_prueba.py

Requiere:
    pip install openpyxl python-dotenv playwright
    playwright install chromium
"""

from __future__ import annotations

import csv
import datetime
import re
import sys
import threading
import time
from datetime import timezone
from pathlib import Path
from typing import Optional

# tkinter es opcional. Import diferido dentro de la función.
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import (
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config  # noqa: E402


# ===========================================================================
# CONSTANTES
# ===========================================================================

NOTIFICATION_TYPE_SELECTORS = [
    "div.notification-type", "div.type", "div.title",
    "div.notification-title", ".notification-header",
]
NOTIFICATION_MESSAGE_SELECTORS = [
    "div.notification-message", "div.message", "div.body",
    "div.notification-body", "div.text", "p.message",
]
NOTIFICATION_CLOSE_SELECTORS = [
    "app-notification-container div.close",
    "app-notification-container button",
    "div.notification-slider div.close",
    "div.notification-slider button",
    "app-notification-slider div.close",
    "app-notification-slider button",
    "app-ui-notifications div.close",
    "app-ui-notifications button",
    "app-ui-notifications div.accept",
    "app-ui-notifications div.ok",
    "div.notifications div.close",
    "div.notifications button",
    "div.notification-panel div.close",
    "button:has-text('Aceptar')",
    "button:has-text('OK')",
    "button:has-text('Cerrar')",
    "div.modal div.close",
    "[role='dialog'] div.close",
    "[role='dialog'] button",
    "[role='alertdialog'] button",
]

HAMBURGER_SELECTORS = [
    "css=div.toggle.menu",
    "css=div.toggle.menu svg",
    "css=div.header div.toggle.menu",
    "css=app-ui-header div.toggle.menu",
    "[aria-label='menu' i]", "[aria-label='menú' i]",
    ".hamburger", ".menu-toggle", ".navbar-toggle", ".burger", ".nav-toggle",
    "header button", "nav button", ".navbar button",
]

LOGIN_SELECTORS = [
    "input[name='username']", "input[name='user']", "input[name='login']",
    "input[id='username']", "input[id='user']",
    "input[type='text'][placeholder*='usu' i]",
    "input[type='text'][placeholder*='user' i]",
    "input[autocomplete='username']",
]
PASSWORD_SELECTORS = [
    "input[name='password']", "input[name='pass']",
    "input[id='password']", "input[id='pass']",
    "input[type='password']", "input[autocomplete='current-password']",
]
SUBMIT_SELECTORS = [
    "button[type='submit']", "input[type='submit']",
    "button:has-text('Ingresar')", "button:has-text('Entrar')",
    "button:has-text('Login')", "button:has-text('Iniciar sesión')",
    "a:has-text('Ingresar')",
]

MAX_FILENAME_LENGTH = 80
MENU_NAVIGATION_DELAY_MS = 600
MENU_REOPEN_DELAY_MS = 500
BROWSER_SLOW_MO_MS = 300
ENTRE_CASOS_PAUSA_MS = 400

# Columnas Excel
COL_CAMINO_HEADER = "Camino a la Pantalla"
COL_TIPO_HEADER = "Tipo Pantalla"
COL_ACCION_HEADER = "Acción"
COL_RESULTADO_HEADER = "Resultado"
COL_DATO_PREFIX = "Dato"

# Tipos de pantalla
TIPO_CONSULTA = "Consulta"
TIPO_CRUD = "CRUD"
TIPO_TRANSACCIONES = "Transacciones"
TIPOS_VALIDOS = {TIPO_CONSULTA, TIPO_CRUD, TIPO_TRANSACCIONES}
TIPO_DEFAULT = TIPO_CONSULTA

# Acciones
ACCION_CREATE = "Create"
ACCION_READ = "Read"
ACCION_UPDATE = "Update"
ACCION_DELETE = "Delete"
ACCIONES_VALIDAS = {ACCION_CREATE, ACCION_READ, ACCION_UPDATE, ACCION_DELETE}
TIPOS_REQUIEREN_ACCION = {TIPO_CRUD, TIPO_TRANSACCIONES}

# Estilos para columna Resultado
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ===========================================================================
# HELPERS — capturas, DOM, clic genérico
# ===========================================================================


def screenshot(page: Page, name: str) -> None:
    ts = datetime.datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", name).strip("_")[:MAX_FILENAME_LENGTH]
    filepath = config.SCREENSHOTS_DIR / f"{ts}_{safe}.png"
    page.screenshot(path=str(filepath))
    print(f"  📸  Captura guardada: {filepath.name}")


def dump_dom(page: Page, name: str = "dom") -> Path:
    ts = datetime.datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", name).strip("_")[:MAX_FILENAME_LENGTH]
    filepath = config.SCREENSHOTS_DIR / f"{ts}_{safe}.html"
    filepath.write_text(page.content(), encoding="utf-8")
    print(f"  🧾  DOM guardado: {filepath.name}")
    return filepath


def element_exists(page: Page, selector: str, timeout: int = 3000) -> bool:
    try:
        page.locator(selector).first.wait_for(state="visible", timeout=timeout)
        return True
    except PlaywrightTimeoutError:
        return False


def wait_and_click(page: Page, selectors: list, description: str,
                   timeout: Optional[int] = None) -> None:
    t = timeout or config.TIMEOUT_MS
    last_error = None
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=t)
            loc.click()
            print(f"  ✅  Clic en '{description}' usando: {sel}")
            return
        except Exception as exc:
            last_error = exc
            continue
    raise RuntimeError(f"No se pudo hacer clic en '{description}'. Último error: {last_error}")


# ===========================================================================
# LOGIN
# ===========================================================================


def handle_login(page: Page) -> None:
    username_visible = any(element_exists(page, s, timeout=3000) for s in LOGIN_SELECTORS)
    if not username_visible:
        print("  ℹ️   No se detectó pantalla de login – sesión activa.")
        return

    print("  🔐  Pantalla de login detectada. Ingresando credenciales…")
    if not config.APP_USERNAME or not config.APP_PASSWORD:
        raise ValueError("APP_USERNAME o APP_PASSWORD no están definidos en .env")

    for sel in LOGIN_SELECTORS:
        if element_exists(page, sel, timeout=2000):
            page.fill(sel, config.APP_USERNAME)
            break
    for sel in PASSWORD_SELECTORS:
        if element_exists(page, sel, timeout=2000):
            page.fill(sel, config.APP_PASSWORD)
            break

    screenshot(page, "01_before_login")
    wait_and_click(page, SUBMIT_SELECTORS, "botón de login")
    page.wait_for_load_state("networkidle", timeout=config.TIMEOUT_MS)
    print("  ✅  Login completado.")
    screenshot(page, "02_after_login")


# ===========================================================================
# MENÚ — apertura / cierre
# ===========================================================================


def abrir_menu_hamburguesa(page: Page) -> None:
    """SIEMPRE hace clic en el hamburguesa, sin chequear si ya está abierto."""
    ultimo_error = None
    for sel in ["css=div.toggle.menu", "css=app-ui-header div.toggle.menu"]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            loc.click()
            print("  🍔  Hamburguesa ≡ clickeado.")
            page.locator("div.ui-menu-item").first.wait_for(state="visible", timeout=5000)
            return
        except Exception as exc:
            ultimo_error = exc
            continue
    try:
        wait_and_click(page, HAMBURGER_SELECTORS, "menú hamburguesa")
        page.locator("div.ui-menu-item").first.wait_for(state="visible", timeout=5000)
    except Exception:
        raise RuntimeError(f"No se pudo hacer clic en el hamburguesa. Último: {ultimo_error}")


def cerrar_menu_si_abierto(page: Page) -> None:
    menu_abierto = False
    for sel in ["div.ui-menu-item", "div.menu-options"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=1000)
            menu_abierto = True
            break
        except PlaywrightTimeoutError:
            continue
    if not menu_abierto:
        return

    print("  🔒  Cerrando menú con clic fuera…")
    try:
        page.mouse.click(800, 400)
        page.wait_for_timeout(500)
    except Exception:
        pass

    try:
        page.locator("div.ui-menu-item").first.wait_for(state="hidden", timeout=2000)
    except PlaywrightTimeoutError:
        try:
            page.locator("div.menu-title.main div.close").first.click(timeout=2000)
            page.wait_for_timeout(400)
        except Exception:
            pass


# ===========================================================================
# NAVEGACIÓN POR RUTA
# ===========================================================================


def parse_camino(camino: str) -> list:
    """'Golf > Consultas > Asientos' → ['Golf','Consultas','Asientos']"""
    if not camino:
        return []
    partes = re.split(r"\s*>\s*", camino.strip())
    return [p for p in partes if p]


def navegar_ruta_completa(page: Page, ruta: list) -> bool:
    try:
        loc = page.locator("div.ui-menu-item", has_text=ruta[0]).first
        loc.wait_for(state="visible", timeout=5000)
        loc.click()
        page.wait_for_timeout(MENU_NAVIGATION_DELAY_MS)

        for rama in ruta[1:-1]:
            loc = page.locator("div.menu-options", has_text=rama).first
            loc.wait_for(state="visible", timeout=5000)
            loc.click()
            page.wait_for_timeout(MENU_NAVIGATION_DELAY_MS)

        if len(ruta) > 1:
            loc = page.locator("div.menu-options", has_text=ruta[-1]).first
            loc.wait_for(state="visible", timeout=5000)
            loc.click()
            page.wait_for_timeout(MENU_NAVIGATION_DELAY_MS)
        return True
    except Exception as exc:
        print(f"  ❌  Error navegando ruta {ruta}: {exc}")
        return False


# ===========================================================================
# NOTIFICACIONES Y CIERRE DE PANTALLA
# ===========================================================================


def is_notification_modal_visible(page: Page):
    """
    Detecta el modal de notificación de la app Thunder/Angular.

    Estructura DOM confirmada (Pantalla_Notificacion_Unimodal.html):
        app-modal
          div.modal-background
            div.modal
              div.header
                div.title          ← texto " Notificaciones "
              app-notification-modal
                app-notification-modal-item
                  div.notification-modal-item
                    div.header > div.header-title  ← "Errores"
                    div.body > div.content > div.text  ← mensaje

    Retorna el locator de app-modal si es visible, None si no.
    Se identifica porque su div.header > div.title contiene "Notificacion".
    """
    try:
        found = page.evaluate("""
            () => {
                // Buscar app-modal cuyo div.header > div.title contenga "Notificac"
                const modals = document.querySelectorAll('app-modal');
                for (const m of modals) {
                    const titleEl = m.querySelector('div.modal div.header div.title');
                    if (titleEl) {
                        const txt = titleEl.innerText.trim();
                        if (txt.includes('Notificac')) return true;
                    }
                }
                return false;
            }
        """)
        if found:
            # Devolver el locator del app-modal que contiene el título de notificación
            return page.locator(
                "app-modal:has(div.modal div.header div.title)"
            ).first
    except Exception:
        pass
    return None


def handle_notification_modal(page: Page, ruta_str: str) -> Optional[str]:
    """
    Detecta, extrae, registra y cierra el modal de notificación.

    Estructura DOM confirmada:
        app-modal > div.modal-background > div.modal
          > div.header > div.title          : "Notificaciones"
          > div.options > div.option.close  : botón X para cerrar el modal
        app-notification-modal-item > div.notification-modal-item
          > div.header > div.header-title   : categoría ("Errores")
          > div.body > div.content > div.text : mensaje detalle
    """
    modal_loc = is_notification_modal_visible(page)
    if modal_loc is None:
        return None

    print("  🔔  Modal de Notificaciones detectado.")
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", ruta_str)[:MAX_FILENAME_LENGTH]
    screenshot(page, f"NOTIF__{safe}")
    dump_dom(page, f"NOTIF_DOM__{safe}")

    # ── Extraer categoría ("Errores") desde div.header-title ──────────────
    tipo = ""
    try:
        # header-title contiene un div.state (ícono) + texto directo
        # Usamos innerText del elemento completo y limpiamos el texto del ícono
        tipo_raw = page.evaluate("""
            () => {
                const item = document.querySelector(
                    'app-notification-modal-item div.notification-modal-item div.header div.header-title'
                );
                if (!item) return '';
                // Clonar para remover el div.state (ícono) y leer solo el texto
                const clone = item.cloneNode(true);
                const state = clone.querySelector('div.state');
                if (state) state.remove();
                return clone.innerText.trim();
            }
        """)
        if tipo_raw:
            tipo = tipo_raw
    except Exception:
        pass

    # ── Extraer mensaje detalle desde div.text ───────────────────────────
    mensaje = ""
    try:
        mensaje_raw = page.evaluate("""
            () => {
                const el = document.querySelector(
                    'app-notification-modal-item div.notification-modal-item '
                    + 'div.body div.content div.text'
                );
                return el ? el.innerText.trim() : '';
            }
        """)
        if mensaje_raw:
            mensaje = mensaje_raw
    except Exception:
        pass

    # ── Fallback: leer texto completo del item ────────────────────────────
    if not tipo and not mensaje:
        try:
            mensaje = page.evaluate("""
                () => {
                    const el = document.querySelector('app-notification-modal-item');
                    return el ? el.innerText.trim() : '(texto no disponible)';
                }
            """) or "(texto no disponible)"
        except Exception:
            mensaje = "(texto no disponible)"

    texto_notif = f"{tipo}: {mensaje}" if tipo else mensaje
    print(f"  🔔  Notificación → tipo='{tipo}' | mensaje='{mensaje}'")

    # ── Registrar en CSV ──────────────────────────────────────────────────
    log_path = config.SCREENSHOTS_DIR / "notificaciones_log.csv"
    ts_iso = datetime.datetime.now(timezone.utc).isoformat()
    escribir_cabecera = not log_path.exists()
    try:
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if escribir_cabecera:
                w.writerow(["timestamp", "ruta_menu", "tipo", "mensaje"])
            w.writerow([ts_iso, ruta_str, tipo, mensaje])
    except Exception as exc:
        print(f"  ⚠️   No se pudo escribir el log: {exc}")

    # ── Cerrar el modal con div.option.close (la X del header del app-modal)
    cerrado = False
    try:
        # Selector exacto confirmado por DOM:
        # app-modal > div.modal-background > div.modal > div.header > div.options > div.option.close
        close_btn = page.locator("app-modal div.modal div.header div.option.close").first
        close_btn.wait_for(state="visible", timeout=3000)
        close_btn.click()
        page.wait_for_timeout(500)
        # Verificar que se cerró
        try:
            page.locator("app-modal div.modal-background").first.wait_for(
                state="hidden", timeout=2000
            )
            print("  ✅  Modal de Notificaciones cerrado con div.option.close.")
            cerrado = True
        except Exception:
            # Puede haber cerrado aunque el wait_for falle — intentar Escape
            pass
    except Exception as e:
        print(f"  ⚠️   Fallo cierre con div.option.close: {e}")

    if not cerrado:
        # Fallback: Escape
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(500)
            page.locator("app-modal div.modal-background").first.wait_for(
                state="hidden", timeout=1500
            )
            print("  ✅  Modal de Notificaciones cerrado con Escape.")
            cerrado = True
        except Exception:
            pass

    if not cerrado:
        print(f"  ⚠️   No se pudo cerrar el modal de Notificaciones.")
        screenshot(page, f"NOTIF_NO_CERRADA__{safe}")

    return texto_notif


def _cerrar_app_modal_si_existe(page: Page) -> bool:
    """
    Detecta y cierra un <app-modal> genérico (confirmación, formulario, etc.).
    NO cierra el modal de Notificaciones — ese lo maneja handle_notification_modal().

    El modal genérico tiene:
        app-modal > div.modal-background > div.modal > div.header > div.title
    donde el título NO contiene "Notificac".
    """
    try:
        page.locator("app-modal div.modal-background").first.wait_for(
            state="visible", timeout=2000
        )
    except Exception:
        return False

    # Verificar que NO sea el modal de Notificaciones
    es_notificacion = page.evaluate("""
        () => {
            const titleEl = document.querySelector(
                'app-modal div.modal div.header div.title'
            );
            if (!titleEl) return false;
            return titleEl.innerText.trim().includes('Notificac');
        }
    """)
    if es_notificacion:
        # No es un modal genérico — dejarlo para handle_notification_modal
        return False

    print("  🪟  app-modal (genérico) detectado.")
    try:
        page.locator("app-modal div.option.close").first.click(timeout=2000)
        page.wait_for_timeout(400)
        return True
    except Exception:
        pass
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(400)
        return True
    except Exception:
        return False


def cerrar_pantalla_layout(page: Page) -> bool:
    _cerrar_app_modal_si_existe(page)
    try:
        close_btn = page.locator("div.tab.activable.active div.close").first
        close_btn.wait_for(state="visible", timeout=5000)
        close_btn.click()
        page.locator("div.tab.activable.active").first.wait_for(state="hidden", timeout=5000)
        page.locator("div.toggle.menu").first.wait_for(state="visible", timeout=5000)
        print("  ✖️   Pantalla cerrada.")
        return True
    except Exception:
        try:
            page.locator("div.tab.activable.active div.close").first.dispatch_event("click")
            page.locator("div.tab.activable.active").first.wait_for(state="hidden", timeout=5000)
            return True
        except Exception:
            return False


# ===========================================================================
# VENTANA "AVANZAR" (Tkinter con fallback a consola)
# ===========================================================================


def _ventana_avanzar_tkinter(timeout_seg: int, titulo_caso: str) -> str:
    import tkinter as tk  # import diferido

    motivo = {"valor": "timeout"}

    root = tk.Tk()
    root.title("Automatización QA — Avanzar")
    root.attributes("-topmost", True)
    root.geometry("420x180+200+200")
    root.resizable(False, False)

    titulo = titulo_caso[:80] if titulo_caso else "Caso de prueba"
    tk.Label(root, text=titulo, font=("Segoe UI", 10, "bold"),
             wraplength=380, justify="center").pack(pady=(15, 5))

    restante = {"seg": timeout_seg}
    lbl_timer = tk.Label(root, text=f"Auto-avanza en {restante['seg']} s",
                         font=("Segoe UI", 9), fg="#555")
    lbl_timer.pack(pady=(0, 10))

    def on_avanzar():
        motivo["valor"] = "avanzar"
        root.destroy()

    def on_cerrar():
        motivo["valor"] = "cerrada"
        root.destroy()

    btn = tk.Button(root, text="Avanzar", width=14, height=2,
                    font=("Segoe UI", 10, "bold"), command=on_avanzar)
    btn.pack(pady=10)

    def tick():
        if not root.winfo_exists():
            return
        restante["seg"] -= 1
        if restante["seg"] <= 0:
            try:
                root.destroy()
            except Exception:
                pass
            return
        lbl_timer.config(text=f"Auto-avanza en {restante['seg']} s")
        root.after(1000, tick)

    root.protocol("WM_DELETE_WINDOW", on_cerrar)
    root.after(1000, tick)
    root.after(timeout_seg * 1000 + 50, lambda: root.winfo_exists() and root.destroy())

    try:
        btn.focus_set()
    except Exception:
        pass

    root.mainloop()
    return motivo["valor"]


def _ventana_avanzar_consola(timeout_seg: int, titulo_caso: str) -> str:
    print("\n" + "─" * 60)
    print(f"  ⏸   Pausa — {titulo_caso[:80]}")
    print(f"        Presiona ENTER para avanzar  (auto-avanza en {timeout_seg}s)")
    print("─" * 60)
    try:
        import msvcrt
        inicio = time.time()
        ultimo = -1
        while True:
            restante = timeout_seg - int(time.time() - inicio)
            if restante <= 0:
                print("\r  ⏱   Tiempo agotado — avanzando…              ")
                return "timeout"
            if restante != ultimo:
                sys.stdout.write(f"\r  ⏱   Auto-avanza en {restante:2d}s  (ENTER)  ")
                sys.stdout.flush()
                ultimo = restante
            if msvcrt.kbhit():
                ch = msvcrt.getch()
                if ch in (b"\r", b"\n", b" "):
                    print("\r  ▶   ENTER recibido — avanzando…              ")
                    return "avanzar"
            time.sleep(0.1)
    except ImportError:
        pass
    try:
        import select
        inicio = time.time()
        ultimo = -1
        while True:
            restante = timeout_seg - int(time.time() - inicio)
            if restante <= 0:
                print("\r  ⏱   Tiempo agotado — avanzando…              ")
                return "timeout"
            if restante != ultimo:
                sys.stdout.write(f"\r  ⏱   Auto-avanza en {restante:2d}s  (ENTER)  ")
                sys.stdout.flush()
                ultimo = restante
            ready, _, _ = select.select([sys.stdin], [], [], 0.1)
            if ready:
                sys.stdin.readline()
                print("\r  ▶   ENTER recibido — avanzando…              ")
                return "avanzar"
    except Exception:
        pass
    time.sleep(timeout_seg)
    return "timeout"


_TKINTER_AVISADO = {"flag": False}


def ventana_avanzar(timeout_seg: int = 20, titulo_caso: str = "") -> str:
    if timeout_seg <= 0:
        timeout_seg = 20
    try:
        return _ventana_avanzar_tkinter(timeout_seg, titulo_caso)
    except ImportError:
        if not _TKINTER_AVISADO["flag"]:
            print("\n  ⚠️   Tkinter no disponible — usando ventana de consola.")
            _TKINTER_AVISADO["flag"] = True
        return _ventana_avanzar_consola(timeout_seg, titulo_caso)
    except Exception as exc:
        print(f"  ⚠️   Tkinter falló ({exc}). Usando consola.")
        return _ventana_avanzar_consola(timeout_seg, titulo_caso)


# ===========================================================================
# EXCEL — lectura/escritura con soporte a fórmulas (VLOOKUP)
# ===========================================================================


def localizar_columnas(ws) -> dict:
    cols = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        cols[str(cell.value).strip()] = cell.column

    faltan = []
    if COL_CAMINO_HEADER not in cols:
        faltan.append(COL_CAMINO_HEADER)
    if COL_RESULTADO_HEADER not in cols:
        faltan.append(COL_RESULTADO_HEADER)
    if faltan:
        raise ValueError(f"El Excel no contiene: {faltan}. Headers: {list(cols.keys())}")
    if COL_TIPO_HEADER not in cols:
        print(f"  ⚠️   Columna '{COL_TIPO_HEADER}' no encontrada — se asumirá '{TIPO_DEFAULT}'.")
    if COL_ACCION_HEADER not in cols:
        print(f"  ⚠️   Columna '{COL_ACCION_HEADER}' no encontrada — CRUD/Transacciones requerirán acción.")
    return cols


def normalizar_tipo(valor) -> str:
    if valor is None:
        return TIPO_DEFAULT
    s = str(valor).strip()
    if s == "":
        return TIPO_DEFAULT
    s_lower = s.lower()
    if s_lower in ("consulta", "consultas", "consult", "c"):
        return TIPO_CONSULTA
    if s_lower in ("crud", "mantenedor", "mantenedores", "abm"):
        return TIPO_CRUD
    if s_lower in ("transacciones", "transaccion", "transacción",
                   "operacion", "operación", "operaciones", "trx", "t"):
        return TIPO_TRANSACCIONES
    return s


def normalizar_accion(valor) -> Optional[str]:
    if valor is None:
        return None
    s = str(valor).strip()
    if s == "":
        return None
    s_lower = s.lower()
    if s_lower in ("create", "crear", "nuevo", "agregar", "add", "alta", "c", "n"):
        return ACCION_CREATE
    if s_lower in ("read", "leer", "consultar", "buscar", "ver", "query", "r", "q"):
        return ACCION_READ
    if s_lower in ("update", "actualizar", "modificar", "editar", "edit", "u", "m", "e"):
        return ACCION_UPDATE
    if s_lower in ("delete", "eliminar", "borrar", "baja", "del", "d", "b"):
        return ACCION_DELETE
    return s


def _resolver_referencia_celda(ref: str, wb, ws_actual):
    if (ref.startswith('"') and ref.endswith('"')) or \
       (ref.startswith("'") and ref.endswith("'") and "!" not in ref):
        return ref[1:-1]
    m = re.match(r"^(?:('[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?\$?([A-Z]+)\$?(\d+)$", ref)
    if not m:
        return None
    sheet_name = m.group(1)
    if sheet_name and sheet_name.startswith("'") and sheet_name.endswith("'"):
        sheet_name = sheet_name[1:-1]
    target_ws = wb[sheet_name] if sheet_name else ws_actual
    from openpyxl.utils import column_index_from_string
    col = column_index_from_string(m.group(2))
    row = int(m.group(3))
    return target_ws.cell(row=row, column=col).value


_VLOOKUP_RE = re.compile(
    r"^=VLOOKUP\(\s*(.+?)\s*,\s*(.+?)\s*,\s*(\d+)"
    r"(?:\s*,\s*(?:TRUE|FALSE|0|1|VERDADERO|FALSO))?\s*\)\s*$",
    re.IGNORECASE,
)
_RANGO_RE = re.compile(
    r"^(?:('[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"\$?([A-Z]+)\$?(\d+)?\s*:\s*\$?([A-Z]+)\$?(\d+)?$"
)


def _evaluar_vlookup(formula: str, wb, ws_actual):
    m = _VLOOKUP_RE.match(formula.strip())
    if not m:
        return None
    lookup_ref, range_ref, col_idx = m.group(1).strip(), m.group(2).strip(), int(m.group(3))

    lookup_value = _resolver_referencia_celda(lookup_ref, wb, ws_actual)
    if lookup_value is None:
        return None

    range_m = _RANGO_RE.match(range_ref)
    if not range_m:
        return None
    sheet_name = range_m.group(1)
    if sheet_name and sheet_name.startswith("'") and sheet_name.endswith("'"):
        sheet_name = sheet_name[1:-1]
    from openpyxl.utils import column_index_from_string
    start_col = column_index_from_string(range_m.group(2))
    start_row = int(range_m.group(3)) if range_m.group(3) else 1
    end_col = column_index_from_string(range_m.group(4))
    end_row_str = range_m.group(5)

    target_ws = wb[sheet_name] if sheet_name else ws_actual
    end_row = int(end_row_str) if end_row_str else target_ws.max_row

    lookup_str = str(lookup_value).strip()
    for row in range(start_row, end_row + 1):
        cell_val = target_ws.cell(row=row, column=start_col).value
        if cell_val is None:
            continue
        if str(cell_val).strip() == lookup_str:
            return_col = start_col + col_idx - 1
            if return_col > end_col:
                return None
            return target_ws.cell(row=row, column=return_col).value
    return None


def leer_celda(ws, ws_valores, fila: int, col: int, wb=None):
    valor = ws.cell(row=fila, column=col).value
    if not (isinstance(valor, str) and valor.startswith("=")):
        return valor

    coord = ws.cell(row=fila, column=col).coordinate

    if ws_valores is not None:
        valor_calc = ws_valores.cell(row=fila, column=col).value
        if valor_calc is not None:
            return valor_calc

    if wb is not None and "VLOOKUP" in valor.upper():
        valor_eval = _evaluar_vlookup(valor, wb, ws)
        if valor_eval is not None:
            print(f"  🧮  Celda {coord}: fórmula evaluada en Python → {valor_eval!r}")
            return valor_eval

    print(f"  ⚠️   Celda {coord} tiene fórmula sin valor calculable:")
    print(f"         {valor}")
    print(f"         Soluciones: (a) abrir en Excel, F9, Ctrl+S")
    print(f"                     (b) verificar que la fórmula sea VLOOKUP exacto")
    return None


def leer_datos_fila(ws, fila: int, columnas: dict, ws_valores=None, wb=None) -> list:
    datos = []
    for n in range(1, 21):
        header = f"{COL_DATO_PREFIX}{n}"
        col = columnas.get(header)
        if col is None:
            continue
        valor = leer_celda(ws, ws_valores, fila, col, wb=wb)
        if valor is None or str(valor).strip() == "":
            continue
        datos.append(str(valor).strip())
    return datos


def escribir_resultado(ws, fila: int, col_resultado: int, texto: str, ok: bool) -> None:
    cell = ws.cell(row=fila, column=col_resultado, value=texto)
    cell.fill = FILL_OK if ok else FILL_ERROR


# ===========================================================================
# HELPERS DE LLENADO DE FORMULARIOS
# ===========================================================================


def parse_dato(s: str) -> tuple:
    """
    Parsea un dato del Excel. Soporta dos formatos:
      [Frame].Field.Valor   → con frame entre corchetes (frame opcional)
      Field.Valor           → sin frame, primer punto separa Field y Valor

    Ejemplos:
      'Código.X'                       → (None, 'Código', 'X')
      'Idioma.es-CL'                   → (None, 'Idioma', 'es-CL')
      'Descripción.CREA QA AUT May 26' → (None, 'Descripción', 'CREA QA AUT May 26')
      '[Datos].RUT.11.111.111-1'       → ('Datos', 'RUT', '11.111.111-1')

    Retorna (frame, field, valor). Si está malformado retorna (None, None, s).
    """
    if not s:
        return (None, None, "")
    s = s.strip()
    # Caso con frame entre corchetes
    if s.startswith("["):
        m = re.match(r"^\[([^\]]+)\]\.([^.]+)\.(.+)$", s)
        if m:
            return (m.group(1), m.group(2), m.group(3))
        return (None, None, s)
    # Caso sin frame: split en el PRIMER punto
    idx = s.find(".")
    if idx == -1:
        return (None, s, "")
    return (None, s[:idx], s[idx + 1:])


def click_by_text(page: Page, text: str, alternativas: Optional[list] = None,
                  timeout: int = 3000) -> bool:
    """
    Hace clic en un elemento clickeable cuyo texto contenga `text`.
    Si falla, prueba con `alternativas` (lista de strings).

    Soporta:
      - <button>, <a>, <input type='button|submit'>
      - Framework Thunder (Angular): <div class="option"> con <span class="label">
      - Componentes Angular Material (app-ui-button, app-button)
      - Elementos con role='button'
    """
    candidatos = [text] + (alternativas or [])
    # Variantes del texto (con y sin "+", para "+ Añadir" vs "Añadir")
    todos = []
    for t in candidatos:
        todos.append(t)
        # Si empieza con "+ ", agregar también la versión sin el "+"
        if t.startswith("+ "):
            todos.append(t[2:].strip())

    for t in todos:
        t_escaped = t.replace("'", "\\'")
        selectores = [
            # === Framework Thunder (Angular custom) ===
            # Botón div.option con span.label conteniendo el texto exacto
            f"div.option:has(span.label:text-is('{t_escaped}'))",
            f"div.option:has(span.label:has-text('{t_escaped}'))",
            f"div.option:has-text('{t_escaped}')",
            # === Botones HTML estándar ===
            f"button:has-text('{t_escaped}')",
            f"a:has-text('{t_escaped}')",
            f"input[type='button'][value='{t_escaped}']",
            f"input[type='submit'][value='{t_escaped}']",
            # === Roles ARIA ===
            f"[role='button']:has-text('{t_escaped}')",
            # === Componentes Angular Material / custom ===
            f"app-ui-button:has-text('{t_escaped}')",
            f"app-button:has-text('{t_escaped}')",
            # === Spans/divs con label exacto (último recurso) ===
            f"span.label:text-is('{t_escaped}')",
            # === Texto exacto en cualquier elemento ===
            f":text-is('{t_escaped}')",
        ]
        for sel in selectores:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=timeout)
                loc.click()
                print(f"        clic OK: '{t}' usando {sel}")
                return True
            except Exception:
                continue
    return False


def fill_field_by_label(page: Page, label_text: str, valor: str,
                        timeout: int = 2000) -> bool:
    """
    Encuentra un input/select/textarea asociado a un label/placeholder/formcontrol
    cuyo nombre coincida con `label_text` y lo llena con `valor`.

    Prueba muchas estrategias en orden de prioridad. La primera familia es
    para Thunder framework (Angular custom usado por la app).
    Devuelve True si pudo llenar el campo.
    """
    label_escaped = label_text.replace("'", "\\'")
    selectores = [
        # === FRAMEWORK THUNDER (Angular custom — prioridad alta) ===
        # Estructura: app-form-text > div.label > label "X"  +  div.input-group > input
        f"app-form-text:has(label:has-text('{label_escaped}')) input",
        f"app-form-text:has(label:has-text('{label_escaped}')) textarea",
        f"app-form-text:has(label:has-text('{label_escaped}')) select",
        f"app-form-field:has(label:has-text('{label_escaped}')) input",
        f"app-form-field:has(label:has-text('{label_escaped}')) textarea",
        f"app-form-field:has(label:has-text('{label_escaped}')) select",
        # === Reactive Forms (Angular estándar) ===
        f"input[formcontrolname='{label_escaped}']",
        f"textarea[formcontrolname='{label_escaped}']",
        f"select[formcontrolname='{label_escaped}']",
        # === Atributos name/id ===
        f"input[name='{label_escaped}']",
        f"input[id='{label_escaped}']",
        f"textarea[name='{label_escaped}']",
        # === Angular Material — mat-form-field ===
        f"mat-form-field:has(mat-label:has-text('{label_escaped}')) input",
        f"mat-form-field:has(mat-label:has-text('{label_escaped}')) textarea",
        f"mat-form-field:has(label:has-text('{label_escaped}')) input",
        # === Componentes Angular UI genéricos ===
        f"app-ui-input:has-text('{label_escaped}') input",
        f"app-ui-textbox:has-text('{label_escaped}') input",
        f"app-ui-textarea:has-text('{label_escaped}') textarea",
        f"app-ui-select:has-text('{label_escaped}') select",
        f"app-ui-select:has-text('{label_escaped}') input",
        # === Label adyacente ===
        f"label:has-text('{label_escaped}') + input",
        f"label:has-text('{label_escaped}') + textarea",
        f"label:has-text('{label_escaped}') + select",
        f"label:has-text('{label_escaped}') ~ input",
        # === Contenedor con label hijo ===
        f"div:has(> label:has-text('{label_escaped}')) input",
        f"div:has(> label:has-text('{label_escaped}')) textarea",
        # === Placeholder ===
        f"input[placeholder*='{label_escaped}' i]",
        f"textarea[placeholder*='{label_escaped}' i]",
        # === aria-label ===
        f"input[aria-label*='{label_escaped}' i]",
        f"textarea[aria-label*='{label_escaped}' i]",
    ]

    for sel in selectores:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=timeout)
            # Detectar si es select nativo o input
            tag = loc.evaluate("el => el.tagName.toLowerCase()")
            if tag == "select":
                # Para <select>, seleccionar por label u option value
                try:
                    loc.select_option(label=valor)
                except Exception:
                    loc.select_option(value=valor)
            else:
                # Limpiar y rellenar
                try:
                    loc.fill("")
                except Exception:
                    pass
                loc.fill(valor)
                # Disparar input + change + blur para que Angular y Thunder
                # ejecuten sus validadores y reactiven el form como "dirty"
                try:
                    loc.evaluate("""el => {
                        el.dispatchEvent(new Event('input', {bubbles: true}));
                        el.dispatchEvent(new Event('change', {bubbles: true}));
                        el.dispatchEvent(new Event('blur', {bubbles: true}));
                    }""")
                except Exception:
                    pass
            print(f"        fill OK: '{label_text}' = '{valor}' usando {sel}")
            return True
        except Exception:
            continue
    return False


# ===========================================================================
# MÓDULOS POR TIPO DE PANTALLA
# ===========================================================================


def procesar_consulta(page: Page, ruta: list, datos: list,
                      accion: Optional[str] = None) -> tuple:
    """Pantalla de Consulta (solo lectura). La acción se ignora."""
    if accion:
        print(f"  ℹ️   [CONSULTA] acción '{accion}' ignorada — Consulta es solo lectura.")
    print(f"  📖  [CONSULTA] pantalla cargada — {len(datos)} filtro(s) en Excel.")
    for d in datos:
        print(f"        filtro pendiente: {d}")
    return True, "OK (Consulta)"


def _cerrar_modal_confirmacion(page: Page) -> bool:
    """
    Cierra el modal de confirmación genérico de Thunder (app-modal que NO es
    el de Notificaciones). Busca el botón 'Aceptar' dentro del modal activo.

    Estructura DOM del modal de confirmación Thunder:
        app-modal > div.modal-background > div.modal
          > div.content > div.body > div.footer > div.options
            > div.option   ← cada opción (Aceptar, Cancelar, etc.)
                > div.label  ← texto del botón

    Retorna True si encontró y clicó el botón, False si no apareció.
    """
    # Selectores en orden de prioridad — todos apuntan al botón Aceptar
    # dentro del modal de confirmación (no del de Notificaciones)
    selectores_aceptar = [
        # Opción Thunder: div.option con div.label conteniendo el texto
        "app-modal div.modal div.footer div.option:has(div.label:has-text('Aceptar'))",
        "app-modal div.modal div.footer div.option:has-text('Aceptar')",
        # Variantes de texto
        "app-modal div.modal div.footer div.option:has(div.label:has-text('Sí'))",
        "app-modal div.modal div.footer div.option:has(div.label:has-text('Si'))",
        "app-modal div.modal div.footer div.option:has(div.label:has-text('OK'))",
        "app-modal div.modal div.footer div.option:has(div.label:has-text('Confirmar'))",
        # Fallback: cualquier botón/div Aceptar visible en la página
        "button:has-text('Aceptar')",
        "button:has-text('Sí')",
    ]
    for sel in selectores_aceptar:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            loc.click()
            page.wait_for_timeout(500)
            print(f"  ✅  Modal de confirmación cerrado con: {sel}")
            return True
        except Exception:
            continue
    return False


def procesar_crud_create(page: Page, ruta: list, datos: list) -> tuple:
    """
    CRUD/Create — flujo:
      1. Clic en botón '+ Añadir'.
      2. Para cada dato del Excel → parsear y llenar el campo.
      3. Clic en botón submit del formulario (guardar).
      4. Modal de confirmación → SIEMPRE aparece → clic en 'Aceptar'.
      5. Verificar si aparece modal de Notificación de error (opcional):
            - Sí → cerrar el modal, registrar en log, retornar ERROR con el mensaje
            - No → éxito confirmado
    """
    ruta_str = " > ".join(ruta)
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", ruta_str)[:MAX_FILENAME_LENGTH]

    # ── 1. Clic en "+ Añadir" ─────────────────────────────────────────────
    print("  🆕  Clic en '+ Añadir'…")
    if not click_by_text(page, "+ Añadir",
                         alternativas=["Añadir", "Nuevo", "Agregar", "+"]):
        screenshot(page, f"ERR_no_anadir__{safe}")
        dump_dom(page, f"ERR_no_anadir__{safe}")
        return False, "FALLO: no se encontró el botón '+ Añadir'"
    page.wait_for_timeout(800)
    screenshot(page, f"form_abierto__{safe}")
    dump_dom(page, f"form_abierto__{safe}")

    # ── 2. Llenar los campos ──────────────────────────────────────────────
    no_llenados = []
    for dato in datos:
        frame, field, valor = parse_dato(dato)
        if field is None:
            print(f"  ⚠️   Dato malformado: {dato!r}")
            no_llenados.append(dato)
            continue
        marca_frame = f"[{frame}] " if frame else ""
        print(f"  ✏️   Llenando {marca_frame}'{field}' = '{valor}'")
        if not fill_field_by_label(page, field, valor):
            print(f"  ❌  No se pudo llenar el campo '{field}'")
            screenshot(page, f"ERR_campo__{field}__{safe}")
            no_llenados.append(dato)
    if no_llenados:
        dump_dom(page, f"form_no_llenado__{safe}")
        return False, f"FALLO: no se llenaron campos: {no_llenados}"

    screenshot(page, f"form_lleno__{safe}")
    dump_dom(page, f"form_lleno__{safe}")

    # ── 3. Clic en botón submit del formulario ────────────────────────────
    # En Thunder: <button type='submit' name='btnSubmit'>
    # DISTINTO al div.option "+ Añadir" de la lista principal.
    print("  💾  Clic en botón guardar del formulario…")
    form_submit_selectors = [
        "app-form-button button[type='submit']",
        "button[name='btnSubmit']",
        "button[type='submit']:has(div.label:has-text('Añadir'))",
        "button[type='submit']:has(div.label:has-text('Crear'))",
        "button[type='submit']:has(div.label:has-text('Guardar'))",
        "button[type='submit']:has-text('Añadir')",
        "button[type='submit']:has-text('Crear')",
        "button[type='submit']",
        "app-form-button button",
    ]
    submit_ok = False
    for sel in form_submit_selectors:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=2000)
            loc.click()
            print(f"        clic OK: guardar usando {sel}")
            submit_ok = True
            break
        except Exception:
            continue
    if not submit_ok:
        screenshot(page, f"ERR_no_submit__{safe}")
        dump_dom(page, f"ERR_no_submit__{safe}")
        return False, "FALLO: no se encontró el botón de guardar del formulario"
    page.wait_for_timeout(600)

    # ── 4. Modal de confirmación — SIEMPRE aparece → clic en 'Aceptar' ───
    print("  🔲  Esperando modal de confirmación…")
    if _cerrar_modal_confirmacion(page):
        print("  ✅  Confirmación aceptada.")
    else:
        # No debería pasar según el comportamiento confirmado, pero si ocurre
        # se registra como advertencia y se continúa para detectar el resultado
        print("  ⚠️   Modal de confirmación no encontrado — continuando igualmente.")
        screenshot(page, f"WARN_sin_confirmacion__{safe}")

    # ── 5. Verificar si aparece modal de Notificación de error (opcional) ─
    # La app muestra este modal SOLO si hubo un error al procesar la operación.
    # Esperar un momento para que Angular procese la respuesta del backend.
    page.wait_for_timeout(1500)
    notif = handle_notification_modal(page, ruta_str)
    if notif:
        # Error de la aplicación — informar en Excel y continuar con el siguiente caso
        print(f"  ❌  Error reportado por la aplicación: {notif}")
        return False, f"ERROR app: {notif}"

    print("  🎉  Sin notificación de error — creación exitosa.")
    return True, "OK (CRUD/Create)"


def _fill_ag_floating_filter_by_col_id(page: Page, col_id: str, valor: str,
                                        timeout: int = 4000) -> bool:
    """
    Llena el filtro flotante de AG-Grid para la columna cuyo col-id coincida.

    La estructura del DOM es:
      .ag-header-row-column-filter
        .ag-floating-filter[aria-colindex=N]   ← columna col-id=<col_id>
          .ag-floating-filter-full-body
            app-default-text-floating-filter (o number)
              input.fr-inp

    Para encontrar el input correcto usamos evaluate() para hacer la búsqueda
    en JS (más robusta con AG-Grid virtual-DOM).
    """
    try:
        input_loc = page.evaluate_handle(
            """(colId) => {
                // Buscar el th del header con ese col-id
                const header = document.querySelector(
                    `.ag-header-cell[col-id="${colId}"]`
                );
                if (!header) return null;

                // La columna de filtros flotantes tiene el mismo aria-colindex
                const colIdx = header.getAttribute('aria-colindex');
                if (!colIdx) return null;

                // Encontrar la celda de filtro con ese aria-colindex en la fila de filtros
                const filterCell = document.querySelector(
                    `.ag-header-row-column-filter .ag-floating-filter[aria-colindex="${colIdx}"]`
                );
                if (!filterCell) return null;

                return filterCell.querySelector('input');
            }""",
            col_id
        )
        if not input_loc:
            return False

        inp = page.locator("css=*").filter(has=input_loc)
        # Más simple: usar evaluate para obtener el índice y luego nth
        idx = page.evaluate(
            """(colId) => {
                const header = document.querySelector(`.ag-header-cell[col-id="${colId}"]`);
                if (!header) return -1;
                const colIdx = header.getAttribute('aria-colindex');
                if (!colIdx) return -1;
                const filterCell = document.querySelector(
                    `.ag-header-row-column-filter .ag-floating-filter[aria-colindex="${colIdx}"]`
                );
                if (!filterCell) return -1;
                const inp = filterCell.querySelector('input');
                if (!inp) return -1;
                // Obtener índice entre todos los inputs del documento
                const allInputs = Array.from(document.querySelectorAll('input'));
                return allInputs.indexOf(inp);
            }""",
            col_id
        )
        if idx < 0:
            return False

        loc = page.locator("input").nth(idx)
        loc.wait_for(state="visible", timeout=timeout)
        loc.click()
        loc.fill("")
        loc.fill(valor)
        loc.press("Enter")
        print(f"        filter OK: col_id='{col_id}' valor='{valor}'")
        return True
    except Exception as exc:
        print(f"        filter ERR: col_id='{col_id}' — {exc}")
        return False


def _fill_ag_floating_filter_by_header_text(page: Page, header_text: str,
                                             valor: str, timeout: int = 4000) -> bool:
    """
    Llena el filtro flotante de AG-Grid buscando la columna por su texto de cabecera.
    Estrategia: encuentra el índice de columna (aria-colindex) del header cuyo
    texto contenga header_text, luego busca el input en la fila de filtros flotantes.
    """
    try:
        idx = page.evaluate(
            """(headerText) => {
                // Buscar entre los ag-header-cell-text el que contenga el texto
                const headers = Array.from(
                    document.querySelectorAll('.ag-header-row-column .ag-header-cell')
                );
                let targetColIdx = null;
                for (const h of headers) {
                    const txt = h.querySelector('.ag-header-cell-text');
                    if (txt && txt.innerText.trim() === headerText) {
                        targetColIdx = h.getAttribute('aria-colindex');
                        break;
                    }
                }
                if (!targetColIdx) {
                    // Intento con contains
                    for (const h of headers) {
                        const txt = h.querySelector('.ag-header-cell-text');
                        if (txt && txt.innerText.trim().toLowerCase()
                                .includes(headerText.toLowerCase())) {
                            targetColIdx = h.getAttribute('aria-colindex');
                            break;
                        }
                    }
                }
                if (!targetColIdx) return -1;

                const filterCell = document.querySelector(
                    `.ag-header-row-column-filter .ag-floating-filter[aria-colindex="${targetColIdx}"]`
                );
                if (!filterCell) return -1;
                const inp = filterCell.querySelector('input');
                if (!inp) return -1;
                const allInputs = Array.from(document.querySelectorAll('input'));
                return allInputs.indexOf(inp);
            }""",
            header_text
        )
        if idx < 0:
            return False

        loc = page.locator("input").nth(idx)
        loc.wait_for(state="visible", timeout=timeout)
        loc.click()
        loc.fill("")
        loc.fill(valor)
        loc.press("Enter")
        print(f"        filter OK: header='{header_text}' valor='{valor}'")
        return True
    except Exception as exc:
        print(f"        filter ERR: header='{header_text}' — {exc}")
        return False


def _ag_grid_get_row_count(page: Page) -> int:
    """
    Devuelve el número de filas visibles en la grilla AG-Grid activa.
    Usa la paginación si está disponible; si no, cuenta los ag-row del DOM.
    """
    try:
        # Intentar leer de la paginación: "Mostrando de X a Y de TOTAL"
        total_text = page.evaluate("""
            () => {
                // Buscar el label de total en la paginación
                const labels = Array.from(
                    document.querySelectorAll('app-pagination .row-viewer .label')
                );
                // El total es el último número en la secuencia
                const texts = labels.map(l => l.innerText.trim()).filter(t => /^\\d+$/.test(t));
                return texts.length > 0 ? parseInt(texts[texts.length - 1]) : -1;
            }
        """)
        if total_text is not None and total_text >= 0:
            return total_text
    except Exception:
        pass

    # Fallback: contar ag-row visibles
    try:
        count = page.evaluate("""
            () => document.querySelectorAll(
                '.ag-center-cols-container .ag-row:not(.ag-hidden)'
            ).length
        """)
        return count or 0
    except Exception:
        return 0


def _ag_grid_get_visible_col_ids(page: Page) -> list:
    """Devuelve la lista de col-id de columnas visibles en la grilla."""
    try:
        return page.evaluate("""
            () => Array.from(
                document.querySelectorAll(
                    '.ag-header-row-column .ag-header-cell[col-id]'
                )
            ).map(h => h.getAttribute('col-id')).filter(Boolean)
        """)
    except Exception:
        return []


def _ag_grid_read_rows(page: Page, max_rows: int = 100) -> list:
    """
    Lee los datos de las filas visibles de la grilla AG-Grid activa usando
    la API interna de AG-Grid (window.agGridApi si está expuesta) o
    leyendo el texto de las celdas del DOM virtualizado.

    Retorna lista de dicts {col_id: valor} para cada fila.
    """
    try:
        rows = page.evaluate("""
            () => {
                // Intentar acceder a la API de AG-Grid
                // AG-Grid expone el api en el elemento ag-grid-angular
                const gridEl = document.querySelector('ag-grid-angular');
                if (!gridEl) return null;

                // En AG-Grid v29+, la API está en gridEl.__agGridAPIRef
                // En versiones anteriores puede estar en gridEl._agGrid
                let api = null;
                if (gridEl.__agGridAPIRef && gridEl.__agGridAPIRef.current) {
                    api = gridEl.__agGridAPIRef.current;
                } else if (gridEl._agGrid) {
                    api = gridEl._agGrid.api;
                }

                if (!api) {
                    // Fallback: buscar en el contexto global
                    // Algunos setups exponen agGrid como variable global
                    return null;
                }

                const rows = [];
                api.forEachNodeAfterFilterAndSort(node => {
                    if (node.data) rows.push(node.data);
                });
                return rows.length > 0 ? rows : null;
            }
        """)
        if rows:
            return rows
    except Exception:
        pass

    # Fallback: leer los innerText de las celdas del DOM
    try:
        rows = page.evaluate("""
            (maxRows) => {
                const container = document.querySelector('.ag-center-cols-container');
                if (!container) return [];

                const result = [];
                const rowEls = Array.from(
                    container.querySelectorAll('[role="row"]')
                ).slice(0, maxRows);

                for (const rowEl of rowEls) {
                    const rowData = {};
                    const cells = rowEl.querySelectorAll('[role="gridcell"][col-id]');
                    for (const cell of cells) {
                        const colId = cell.getAttribute('col-id');
                        rowData[colId] = cell.innerText.trim();
                    }
                    if (Object.keys(rowData).length > 0) {
                        result.push(rowData);
                    }
                }
                return result;
            }
        """, max_rows)
        return rows or []
    except Exception:
        return []


def procesar_crud_read(page: Page, ruta: list, datos: list) -> tuple:
    """
    CRUD/Read — flujo para pantallas con grilla AG-Grid (framework Thunder/Angular):

    La pantalla es una tabla AG-Grid con filtros flotantes por columna.
    El usuario filtra escribiendo en la celda de filtro de la columna y pulsando Enter.

    Flujo:
      1. Para cada dato del Excel → parsear campo y valor.
      2. Identificar el filtro flotante de la columna (por col-id o texto de cabecera).
      3. Llenar el filtro con el valor y presionar Enter.
      4. Esperar a que la grilla se actualice.
      5. Contar filas resultantes:
           - 0 filas → NO ENCONTRADO
           - ≥1 fila  → ENCONTRADO (informar cuántas filas y primer valor)
      6. Retornar resultado y escribir en Excel.

    Formato de datos en Excel:
      - Sin frame: "Código.X"   → filtra columna "Código" con valor "X"
      - Con frame:  "[Grid].Código.X" → ídem (frame se ignora para Read)
    """
    ruta_str = " > ".join(ruta)
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", ruta_str)[:MAX_FILENAME_LENGTH]

    if not datos:
        return False, "ERROR_CONFIG: CRUD/Read requiere al menos un dato (campo de búsqueda)"

    # Esperar a que la grilla esté lista
    try:
        page.locator(".ag-root-wrapper").first.wait_for(state="visible", timeout=5000)
        page.wait_for_timeout(500)
    except PlaywrightTimeoutError:
        return False, "FALLO: No se encontró la grilla AG-Grid en la pantalla"

    # Leer columnas disponibles para logging
    col_ids = _ag_grid_get_visible_col_ids(page)
    print(f"  📊  Columnas grilla: {col_ids}")

    screenshot(page, f"crud_read_inicio__{safe}")

    filtros_aplicados = []
    filtros_fallidos = []

    for dato in datos:
        _frame, field, valor = parse_dato(dato)
        if field is None:
            print(f"  ⚠️   Dato malformado: {dato!r}")
            filtros_fallidos.append(dato)
            continue

        print(f"  🔍  Aplicando filtro: '{field}' = '{valor}'")

        # Estrategia 1: buscar por col-id exacto (ej: "codigo", "descripcion")
        col_id_candidato = field.lower().replace(" ", "_")
        ok = False

        # Primero intentar por col-id exacto
        if col_id_candidato in (c.lower() for c in col_ids):
            col_id_real = next(c for c in col_ids if c.lower() == col_id_candidato)
            ok = _fill_ag_floating_filter_by_col_id(page, col_id_real, valor)

        # Si no, intentar por texto de cabecera (ej: "Código", "Descripción")
        if not ok:
            ok = _fill_ag_floating_filter_by_header_text(page, field, valor)

        # Si tampoco, intentar col-id sin acentos ni mayúsculas
        if not ok:
            for c in col_ids:
                if field.lower() in c.lower() or c.lower() in field.lower():
                    ok = _fill_ag_floating_filter_by_col_id(page, c, valor)
                    if ok:
                        break

        if ok:
            filtros_aplicados.append(f"{field}={valor}")
        else:
            print(f"  ❌  No se pudo aplicar filtro para campo '{field}'")
            filtros_fallidos.append(dato)

    if not filtros_aplicados:
        dump_dom(page, f"crud_read_sin_filtros__{safe}")
        return False, f"FALLO: no se pudo aplicar ningún filtro. Datos: {datos}"

    # Esperar a que la grilla actualice los resultados
    page.wait_for_timeout(1200)
    screenshot(page, f"crud_read_filtrado__{safe}")

    # Contar filas resultantes
    total_filas = _ag_grid_get_row_count(page)
    print(f"  📈  Filas resultantes: {total_filas}")

    # Intentar leer primera fila para incluir en el resultado
    primera_fila_str = ""
    filas = _ag_grid_read_rows(page, max_rows=3)
    if filas:
        primera = filas[0]
        primera_fila_str = " | ".join(
            f"{k}={v}" for k, v in primera.items() if v
        )

    filtros_str = ", ".join(filtros_aplicados)

    if total_filas == 0:
        resultado = f"NO ENCONTRADO ({filtros_str})"
        dump_dom(page, f"crud_read_no_encontrado__{safe}")
        return False, resultado
    else:
        detalle = f" → {primera_fila_str}" if primera_fila_str else ""
        resultado = f"ENCONTRADO: {total_filas} fila(s) ({filtros_str}){detalle}"
        return True, resultado


def procesar_crud(page: Page, ruta: list, datos: list,
                  accion: Optional[str] = None) -> tuple:
    """Pantalla de Mantenedor CRUD. Despacha según la acción."""
    if accion is None:
        return False, "ERROR_CONFIG: Acción requerida para CRUD (Create/Read/Update/Delete)"
    if accion not in ACCIONES_VALIDAS:
        return False, f"ERROR_CONFIG: Acción '{accion}' no válida para CRUD"

    print(f"  🗂   [CRUD/{accion}] pantalla cargada — {len(datos)} dato(s).")

    if accion == ACCION_CREATE:
        return procesar_crud_create(page, ruta, datos)

    if accion == ACCION_READ:
        return procesar_crud_read(page, ruta, datos)

    # Update y Delete siguen como stub
    if accion == ACCION_UPDATE:
        print(f"        TODO: localizar registro, editar campos, grabar.")
    elif accion == ACCION_DELETE:
        print(f"        TODO: localizar registro, clic en eliminar, confirmar.")
    for d in datos:
        print(f"        dato pendiente: {d}")
    return True, f"OK (CRUD/{accion} - stub)"


def procesar_transaccion(page: Page, ruta: list, datos: list,
                         accion: Optional[str] = None) -> tuple:
    """Pantalla de Transacciones (operaciones). Requiere acción."""
    if accion is None:
        return False, "ERROR_CONFIG: Acción requerida para Transacciones (Create/Read/Update/Delete)"
    if accion not in ACCIONES_VALIDAS:
        return False, f"ERROR_CONFIG: Acción '{accion}' no válida para Transacciones"

    print(f"  💼  [TRANSACCION/{accion}] pantalla cargada — {len(datos)} dato(s).")
    if accion == ACCION_CREATE:
        print(f"        TODO: clic '+', llenar frames, grabar/procesar.")
    elif accion == ACCION_READ:
        print(f"        TODO: localizar transacción existente.")
    elif accion == ACCION_UPDATE:
        print(f"        TODO: localizar transacción, modificar campos, grabar.")
    elif accion == ACCION_DELETE:
        print(f"        TODO: localizar transacción, anular/eliminar.")
    for d in datos:
        print(f"        dato pendiente: {d}")
    return True, f"OK (Transacción/{accion} - stub)"


DISPATCH_POR_TIPO = {
    TIPO_CONSULTA: procesar_consulta,
    TIPO_CRUD: procesar_crud,
    TIPO_TRANSACCIONES: procesar_transaccion,
}


# ===========================================================================
# PREPARAR ESTADO LIMPIO ENTRE CASOS
# ===========================================================================


def preparar_estado_limpio(page: Page) -> None:
    _cerrar_app_modal_si_existe(page)
    try:
        if is_notification_modal_visible(page) is not None:
            print("  🧹  Notificación residual — cerrándola…")
            handle_notification_modal(page, "_residual_")
    except Exception:
        pass
    try:
        page.locator("div.tab.activable.active").first.wait_for(
            state="visible", timeout=1000
        )
        print("  🧹  Pantalla previa aún abierta — cerrándola…")
        cerrar_pantalla_layout(page)
    except PlaywrightTimeoutError:
        pass
    except Exception:
        pass
    cerrar_menu_si_abierto(page)
    page.wait_for_timeout(ENTRE_CASOS_PAUSA_MS)


# ===========================================================================
# PROCESAR CASO INDIVIDUAL
# ===========================================================================


def procesar_caso(page: Page, ruta: list, datos: list,
                  tipo: str = TIPO_DEFAULT,
                  accion: Optional[str] = None) -> tuple:
    ruta_str = " > ".join(ruta)

    preparar_estado_limpio(page)

    try:
        page.locator("div.toggle.menu").first.wait_for(state="visible", timeout=5000)
    except PlaywrightTimeoutError:
        return False, "FALLO_NAVEGACION: hamburguesa ≡ no visible"

    try:
        abrir_menu_hamburguesa(page)
    except Exception as exc:
        return False, f"FALLO_NAVEGACION: no se pudo abrir el menú ({exc})"

    if not navegar_ruta_completa(page, ruta):
        return False, "FALLO_NAVEGACION: no se pudo recorrer la ruta completa"

    try:
        page.locator("div.tab.activable.active").first.wait_for(
            state="visible", timeout=8000
        )
    except PlaywrightTimeoutError:
        notif = handle_notification_modal(page, ruta_str)
        if notif:
            return False, f"ERROR: {notif}"
        return False, "FALLO_NAVEGACION: la pantalla nunca apareció"

    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", ruta_str)[:MAX_FILENAME_LENGTH]
    screenshot(page, f"caso__{safe}")
    dump_dom(page, f"pantalla_inicial__{safe}")

    notif = handle_notification_modal(page, ruta_str)
    if notif:
        try:
            cerrar_pantalla_layout(page)
        except Exception:
            pass
        return False, f"ERROR: {notif}"

    handler = DISPATCH_POR_TIPO.get(tipo)
    if handler is None:
        try:
            cerrar_pantalla_layout(page)
        except Exception:
            pass
        return False, f"ERROR_CONFIG: Tipo Pantalla '{tipo}' no válido"

    accion_log = f"/{accion}" if accion else ""
    print(f"  ▶  Dispatch → módulo '{tipo}{accion_log}'")
    try:
        ok, resultado = handler(page, ruta, datos, accion=accion)
    except Exception as exc:
        ok, resultado = False, f"EXCEPCION en módulo {tipo}: {exc}"
        try:
            screenshot(page, f"EXC_{tipo}_{safe}")
        except Exception:
            pass

    cerrar_pantalla_layout(page)
    return ok, resultado


# ===========================================================================
# FLUJO PRINCIPAL
# ===========================================================================


def run() -> None:
    print("\n" + "=" * 60)
    print("  Automatización QA — Ejecución de Casos de Prueba")
    print("=" * 60)
    print(f"  URL base       : {config.BASE_URL}")
    print(f"  Headless       : {config.HEADLESS}")
    print(f"  Timeout        : {config.TIMEOUT_MS} ms")
    print(f"  Capturas       : {config.SCREENSHOTS_DIR}")
    print(f"  Excel casos    : {config.EXCEL_CASOS}")
    print(f"  ACTIVA_VENTANA : {config.ACTIVA_VENTANA}")
    print(f"  Tiempo ventana : {config.VENTANA_TIMEOUT_SEG} s")
    print("=" * 60 + "\n")

    if not config.EXCEL_CASOS.exists():
        print(f"❌  No se encontró el Excel: {config.EXCEL_CASOS}")
        sys.exit(1)

    print(f"▶ Abriendo Excel: {config.EXCEL_CASOS.name}")
    wb = load_workbook(config.EXCEL_CASOS)
    ws = wb.active
    try:
        wb_valores = load_workbook(config.EXCEL_CASOS, data_only=True)
        ws_valores = wb_valores[ws.title]
        print(f"  ✅  Modo dual: data_only=True para fórmulas")
    except Exception as exc:
        print(f"  ⚠️   No se pudo abrir en modo data_only: {exc}")
        wb_valores = None
        ws_valores = None

    columnas = localizar_columnas(ws)
    col_camino = columnas[COL_CAMINO_HEADER]
    col_resultado = columnas[COL_RESULTADO_HEADER]
    col_tipo = columnas.get(COL_TIPO_HEADER)
    col_accion = columnas.get(COL_ACCION_HEADER)
    total_filas = ws.max_row - 1
    print(f"  ✅  Hoja: '{ws.title}' — {total_filas} caso(s) detectado(s).\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=config.HEADLESS,
            slow_mo=BROWSER_SLOW_MO_MS,
            args=["--start-maximized"],
        )
        context = browser.new_context(viewport=None, ignore_https_errors=True)
        context.set_default_timeout(config.TIMEOUT_MS)
        page = context.new_page()

        exitosos = []
        fallidos = []

        try:
            print("▶ Cargando URL y login…")
            page.goto(config.BASE_URL, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle", timeout=config.TIMEOUT_MS)
            handle_login(page)
            print()

            for fila in range(2, ws.max_row + 1):
                camino = leer_celda(ws, ws_valores, fila, col_camino, wb=wb)
                if camino is None or str(camino).strip() == "":
                    print(f"[Fila {fila}] (vacía) — se omite")
                    continue

                ruta = parse_camino(str(camino))
                ruta_str = " > ".join(ruta)

                if col_tipo is not None:
                    tipo_raw = leer_celda(ws, ws_valores, fila, col_tipo, wb=wb)
                    tipo = normalizar_tipo(tipo_raw)
                else:
                    tipo = TIPO_DEFAULT

                accion = None
                if col_accion is not None:
                    accion_raw = leer_celda(ws, ws_valores, fila, col_accion, wb=wb)
                    accion = normalizar_accion(accion_raw)

                etiqueta = f"/{accion}" if accion else ""
                print(f"\n{'─'*60}")
                print(f"[Fila {fila}] [{tipo}{etiqueta}] {ruta_str}")
                print(f"{'─'*60}")

                datos = leer_datos_fila(ws, fila, columnas,
                                        ws_valores=ws_valores, wb=wb)

                try:
                    ok, resultado = procesar_caso(page, ruta, datos,
                                                  tipo=tipo, accion=accion)
                except Exception as exc:
                    ok, resultado = False, f"EXCEPCION: {exc}"
                except Exception as exc:
                    ok, resultado = False, f"EXCEPCION: {exc}"
                    try:
                        screenshot(page, f"EXC_fila{fila}")
                    except Exception:
                        pass
                    try:
                        preparar_estado_limpio(page)
                    except Exception:
                        pass

                print(f"  → {resultado}")
                escribir_resultado(ws, fila, col_resultado, resultado, ok)

                try:
                    wb.save(config.EXCEL_CASOS)
                except PermissionError:
                    print(f"  ⚠️   No se pudo guardar el Excel (¿abierto en Excel?).")

                (exitosos if ok else fallidos).append(ruta_str)

                if config.ACTIVA_VENTANA:
                    motivo = ventana_avanzar(
                        timeout_seg=config.VENTANA_TIMEOUT_SEG,
                        titulo_caso=ruta_str,
                    )
                    print(f"  ▶   Continuando ({motivo}).")

            print(f"\n{'='*60}")
            print(f"  ✅  Casos exitosos : {len(exitosos)}")
            print(f"  ❌  Casos fallidos : {len(fallidos)}")
            print(f"{'='*60}")

        except Exception as exc:
            print(f"\n❌  Error fatal: {exc}", file=sys.stderr)
            try:
                screenshot(page, "FATAL_ERROR")
            except Exception:
                pass
            raise
        finally:
            try:
                wb.save(config.EXCEL_CASOS)
                print(f"\n💾  Excel guardado: {config.EXCEL_CASOS}")
            except Exception as exc:
                print(f"\n⚠️   No se pudo guardar el Excel al cerrar: {exc}")
            wb.close()
            if wb_valores is not None:
                wb_valores.close()
            context.close()
            browser.close()


if __name__ == "__main__":
    run()
