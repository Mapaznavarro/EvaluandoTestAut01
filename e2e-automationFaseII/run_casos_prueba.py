"""
run_casos_prueba.py
-------------------
Ejecuta los Casos de Prueba con Datos de Entrada definidos en un Excel.

PRIMERA SOLICITUD (este programa):
    1. Abre la URL y hace login.
    2. Lee el Excel definido por EXCEL_CASOS en .env.
    3. Para cada fila:
         a) Lee la columna "Camino a la Pantalla".
         b) Navega por el menú hasta la pantalla.
         c) Si ACTIVA_VENTANA=Si, abre una ventana con un botón "Avanzar"
            que se autocierra después de VENTANA_TIMEOUT_SEG segundos
            (default 20).
         d) Si la pantalla muestra una notificación de error, escribe
            ese mensaje en la columna "Resultado".
            Si todo OK, escribe "OK".
         e) Cierra la pantalla y pasa al siguiente caso (siempre partiendo
            de nuevo desde el menú hamburguesa).
    4. Al terminar guarda y cierra el Excel.

PASOS FUTUROS (NO implementados aquí, pero sí considerados en el diseño):
    - Usar las columnas Dato1..Dato20 con formato [Frame].Field.Valor
      para rellenar la pantalla y grabar.
    - Si los datos son rechazados → registrar mensaje + cancelar.
    - Si OK → registrar "EXITO".
    El gancho está en la función `procesar_caso()` → ver TODO marcado.

Uso:
    python run_casos_prueba.py

Requiere:
    - .env con APP_USERNAME, APP_PASSWORD, BASE_URL,
      ACTIVA_VENTANA, EXCEL_CASOS, VENTANA_TIMEOUT_SEG.
    - Playwright instalado:  playwright install chromium
    - openpyxl:              pip install openpyxl python-dotenv playwright
"""

from __future__ import annotations

import csv
import datetime
import re
import sys
import threading
import time
from datetime import timezone
from pathlib import Path
from typing import Optional

# tkinter es opcional: si la instalación de Python no lo incluye
# (suele pasar con Python del Microsoft Store o instalaciones minimal),
# usamos un fallback en consola.  Por eso NO se importa a nivel de módulo.
# El import se hace dentro de ventana_avanzar() de forma diferida.

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import (
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config  # noqa: E402


# ===========================================================================
# CONSTANTES
# ===========================================================================

NOTIFICATION_TYPE_SELECTORS = [
    "div.notification-type", "div.type", "div.title",
    "div.notification-title", ".notification-header",
]
NOTIFICATION_MESSAGE_SELECTORS = [
    "div.notification-message", "div.message", "div.body",
    "div.notification-body", "div.text", "p.message",
]
NOTIFICATION_CLOSE_SELECTORS = [
    "app-notification-container div.close",
    "app-notification-container button",
    "div.notification-slider div.close",
    "div.notification-slider button",
    "app-notification-slider div.close",
    "app-notification-slider button",
    "app-ui-notifications div.close",
    "app-ui-notifications button",
    "app-ui-notifications div.accept",
    "app-ui-notifications div.ok",
    "div.notifications div.close",
    "div.notifications button",
    "div.notification-panel div.close",
    "button:has-text('Aceptar')",
    "button:has-text('OK')",
    "button:has-text('Cerrar')",
    "div.modal div.close",
    "[role='dialog'] div.close",
    "[role='dialog'] button",
    "[role='alertdialog'] button",
]

HAMBURGER_SELECTORS = [
    "css=div.toggle.menu",
    "css=div.toggle.menu svg",
    "css=div.header div.toggle.menu",
    "css=app-ui-header div.toggle.menu",
    "[aria-label='menu' i]", "[aria-label='menú' i]",
    ".hamburger", ".menu-toggle", ".navbar-toggle", ".burger", ".nav-toggle",
    "header button", "nav button", ".navbar button",
]

LOGIN_SELECTORS = [
    "input[name='username']", "input[name='user']", "input[name='login']",
    "input[id='username']", "input[id='user']",
    "input[type='text'][placeholder*='usu' i]",
    "input[type='text'][placeholder*='user' i]",
    "input[autocomplete='username']",
]
PASSWORD_SELECTORS = [
    "input[name='password']", "input[name='pass']",
    "input[id='password']", "input[id='pass']",
    "input[type='password']", "input[autocomplete='current-password']",
]
SUBMIT_SELECTORS = [
    "button[type='submit']", "input[type='submit']",
    "button:has-text('Ingresar')", "button:has-text('Entrar')",
    "button:has-text('Login')", "button:has-text('Iniciar sesión')",
    "a:has-text('Ingresar')",
]

MAX_FILENAME_LENGTH = 80
MENU_NAVIGATION_DELAY_MS = 600
MENU_REOPEN_DELAY_MS = 500
BROWSER_SLOW_MO_MS = 300
ENTRE_CASOS_PAUSA_MS = 400  # pausa breve entre el cierre de un caso y el siguiente


# ===========================================================================
# HELPERS — capturas, DOM, clic genérico
# ===========================================================================


def screenshot(page: Page, name: str) -> None:
    ts = datetime.datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", name).strip("_")[:MAX_FILENAME_LENGTH]
    filepath = config.SCREENSHOTS_DIR / f"{ts}_{safe}.png"
    page.screenshot(path=str(filepath))
    print(f"  📸  Captura guardada: {filepath.name}")


def dump_dom(page: Page, name: str = "dom") -> Path:
    ts = datetime.datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", name).strip("_")[:MAX_FILENAME_LENGTH]
    filepath = config.SCREENSHOTS_DIR / f"{ts}_{safe}.html"
    filepath.write_text(page.content(), encoding="utf-8")
    print(f"  🧾  DOM guardado: {filepath.name}")
    return filepath


def element_exists(page: Page, selector: str, timeout: int = 3000) -> bool:
    try:
        page.locator(selector).first.wait_for(state="visible", timeout=timeout)
        return True
    except PlaywrightTimeoutError:
        return False


def wait_and_click(page: Page, selectors: list[str], description: str,
                   timeout: Optional[int] = None) -> None:
    t = timeout or config.TIMEOUT_MS
    last_error: Optional[Exception] = None
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=t)
            loc.click()
            print(f"  ✅  Clic en '{description}' usando: {sel}")
            return
        except Exception as exc:
            last_error = exc
            continue
    raise RuntimeError(
        f"No se pudo hacer clic en '{description}'. "
        f"Último error: {last_error}"
    )


# ===========================================================================
# LOGIN
# ===========================================================================


def handle_login(page: Page) -> None:
    username_visible = any(element_exists(page, s, timeout=3000) for s in LOGIN_SELECTORS)
    if not username_visible:
        print("  ℹ️   No se detectó pantalla de login – sesión activa.")
        return

    print("  🔐  Pantalla de login detectada. Ingresando credenciales…")
    if not config.APP_USERNAME or not config.APP_PASSWORD:
        raise ValueError(
            "APP_USERNAME o APP_PASSWORD no están definidos en .env"
        )

    for sel in LOGIN_SELECTORS:
        if element_exists(page, sel, timeout=2000):
            page.fill(sel, config.APP_USERNAME)
            break
    for sel in PASSWORD_SELECTORS:
        if element_exists(page, sel, timeout=2000):
            page.fill(sel, config.APP_PASSWORD)
            break

    screenshot(page, "01_before_login")
    wait_and_click(page, SUBMIT_SELECTORS, "botón de login")
    page.wait_for_load_state("networkidle", timeout=config.TIMEOUT_MS)
    print("  ✅  Login completado.")
    screenshot(page, "02_after_login")


# ===========================================================================
# MENÚ — apertura / cierre
# ===========================================================================


def abrir_menu_hamburguesa(page: Page) -> None:
    """
    SIEMPRE hace clic en el hamburguesa ≡ y espera a que div.ui-menu-item
    sea visible. Esta función NO chequea si el menú "ya está abierto" —
    entre casos consecutivos esa detección puede dar falso positivo y
    romper la navegación. Cada caso parte desde el hamburguesa.
    """
    ultimo_error: Optional[Exception] = None
    for sel in ["css=div.toggle.menu", "css=app-ui-header div.toggle.menu"]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            loc.click()
            print("  🍔  Hamburguesa ≡ clickeado.")
            page.locator("div.ui-menu-item").first.wait_for(state="visible", timeout=5000)
            return
        except Exception as exc:
            ultimo_error = exc
            continue
    # Último recurso: probar todos los selectores genéricos
    try:
        wait_and_click(page, HAMBURGER_SELECTORS, "menú hamburguesa")
        page.locator("div.ui-menu-item").first.wait_for(state="visible", timeout=5000)
    except Exception:
        raise RuntimeError(
            f"No se pudo hacer clic en el hamburguesa ≡. "
            f"Último error: {ultimo_error}"
        )


def cerrar_menu_si_abierto(page: Page) -> None:
    menu_abierto = False
    for sel in ["div.ui-menu-item", "div.menu-options"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=1000)
            menu_abierto = True
            break
        except PlaywrightTimeoutError:
            continue
    if not menu_abierto:
        return

    print("  🔒  Cerrando menú con clic fuera…")
    try:
        page.mouse.click(800, 400)
        page.wait_for_timeout(500)
    except Exception:
        pass

    try:
        page.locator("div.ui-menu-item").first.wait_for(state="hidden", timeout=2000)
    except PlaywrightTimeoutError:
        # Fallback: botón X del panel raíz
        try:
            page.locator("div.menu-title.main div.close").first.click(timeout=2000)
            page.wait_for_timeout(400)
        except Exception:
            pass


# ===========================================================================
# NAVEGACIÓN POR RUTA
# ===========================================================================


def parse_camino(camino: str) -> list[str]:
    """
    'Golf > Consultas > Contabilidad > Asientos manuales'
      → ['Golf', 'Consultas', 'Contabilidad', 'Asientos manuales']
    """
    if not camino:
        return []
    partes = re.split(r"\s*>\s*", camino.strip())
    return [p for p in partes if p]


def navegar_ruta_completa(page: Page, ruta: list[str]) -> bool:
    """
    Navega toda la ruta desde el menú ya abierto:
      ruta[0]    → div.ui-menu-item
      ruta[1:-1] → div.menu-options (ramas)
      ruta[-1]   → div.menu-options (hoja)
    """
    try:
        loc = page.locator("div.ui-menu-item", has_text=ruta[0]).first
        loc.wait_for(state="visible", timeout=5000)
        loc.click()
        page.wait_for_timeout(MENU_NAVIGATION_DELAY_MS)

        for rama in ruta[1:-1]:
            loc = page.locator("div.menu-options", has_text=rama).first
            loc.wait_for(state="visible", timeout=5000)
            loc.click()
            page.wait_for_timeout(MENU_NAVIGATION_DELAY_MS)

        if len(ruta) > 1:
            loc = page.locator("div.menu-options", has_text=ruta[-1]).first
            loc.wait_for(state="visible", timeout=5000)
            loc.click()
            page.wait_for_timeout(MENU_NAVIGATION_DELAY_MS)

        return True
    except Exception as exc:
        print(f"  ❌  Error navegando ruta {ruta}: {exc}")
        return False


# ===========================================================================
# NOTIFICACIONES Y CIERRE DE PANTALLA
# ===========================================================================


def is_notification_modal_visible(page: Page):
    """
    Detecta la ventana unimodal de notificación buscando un div.title
    cuyo innerText sea exactamente ' Notificaciones '.
    """
    try:
        index = page.evaluate("""
            () => {
                const divs = document.querySelectorAll('div.title');
                for (let i = 0; i < divs.length; i++) {
                    if (divs[i].innerText === ' Notificaciones ') return i;
                }
                return -1;
            }
        """)
        if index >= 0:
            return page.locator("div.title").nth(index)
    except Exception:
        pass
    return None


def handle_notification_modal(page: Page, ruta_str: str) -> Optional[str]:
    """
    Si hay notificación visible: extrae tipo + mensaje, registra en CSV,
    cierra el modal y retorna "{tipo}: {mensaje}". Si no hay, retorna None.
    """
    modal_loc = is_notification_modal_visible(page)
    if modal_loc is None:
        return None

    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", ruta_str)[:MAX_FILENAME_LENGTH]
    screenshot(page, f"NOTIF__{safe}")
    dump_dom(page, f"NOTIF_DOM__{safe}")

    tipo = ""
    for sel in NOTIFICATION_TYPE_SELECTORS:
        try:
            t = modal_loc.locator(sel).first.inner_text().strip()
            if t:
                tipo = t
                break
        except Exception:
            continue

    mensaje = ""
    for sel in NOTIFICATION_MESSAGE_SELECTORS:
        try:
            m = modal_loc.locator(sel).first.inner_text().strip()
            if m:
                mensaje = m
                break
        except Exception:
            continue

    if not tipo and not mensaje:
        try:
            mensaje = modal_loc.inner_text().strip()
        except Exception:
            mensaje = "(texto no disponible)"

    texto_notif = f"{tipo}: {mensaje}" if tipo else mensaje

    # Log a CSV
    log_path = config.SCREENSHOTS_DIR / "notificaciones_log.csv"
    ts_iso = datetime.datetime.now(timezone.utc).isoformat()
    escribir_cabecera = not log_path.exists()
    try:
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if escribir_cabecera:
                w.writerow(["timestamp", "ruta_menu", "tipo", "mensaje"])
            w.writerow([ts_iso, ruta_str, tipo, mensaje])
    except Exception as exc:
        print(f"  ⚠️   No se pudo escribir el log: {exc}")

    # Cerrar
    for sel in NOTIFICATION_CLOSE_SELECTORS:
        try:
            btn = page.locator(sel).first
            btn.wait_for(state="visible", timeout=2000)
            btn.click()
            page.wait_for_timeout(400)
            print(f"  ✅  Notificación cerrada.")
            break
        except Exception:
            continue

    return texto_notif


def _cerrar_app_modal_si_existe(page: Page) -> bool:
    try:
        page.locator("app-modal div.modal-background").first.wait_for(
            state="visible", timeout=2000
        )
    except Exception:
        return False

    print("  🪟  app-modal detectado.")
    try:
        page.locator("app-modal div.option.close").first.click(timeout=2000)
        page.wait_for_timeout(400)
        return True
    except Exception:
        pass
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(400)
        return True
    except Exception:
        return False


def cerrar_pantalla_layout(page: Page) -> bool:
    _cerrar_app_modal_si_existe(page)
    try:
        close_btn = page.locator("div.tab.activable.active div.close").first
        close_btn.wait_for(state="visible", timeout=5000)
        close_btn.click()
        page.locator("div.tab.activable.active").first.wait_for(state="hidden", timeout=5000)
        page.locator("div.toggle.menu").first.wait_for(state="visible", timeout=5000)
        print("  ✖️   Pantalla cerrada.")
        return True
    except Exception:
        try:
            page.locator("div.tab.activable.active div.close").first.dispatch_event("click")
            page.locator("div.tab.activable.active").first.wait_for(state="hidden", timeout=5000)
            return True
        except Exception:
            return False


# ===========================================================================
# VENTANA "AVANZAR"  (Tkinter con fallback a consola)
# ===========================================================================


def _ventana_avanzar_tkinter(timeout_seg: int, titulo_caso: str) -> str:
    """
    Implementación con Tkinter (ventana gráfica).
    Lanza ImportError si tkinter no está disponible.
    """
    import tkinter as tk  # diferido

    motivo = {"valor": "timeout"}

    root = tk.Tk()
    root.title("Automatización QA — Avanzar")
    root.attributes("-topmost", True)
    root.geometry("420x180+200+200")
    root.resizable(False, False)

    titulo = titulo_caso[:80] if titulo_caso else "Caso de prueba"
    tk.Label(
        root, text=titulo, font=("Segoe UI", 10, "bold"),
        wraplength=380, justify="center",
    ).pack(pady=(15, 5))

    restante = {"seg": timeout_seg}
    lbl_timer = tk.Label(
        root, text=f"Auto-avanza en {restante['seg']} s",
        font=("Segoe UI", 9), fg="#555",
    )
    lbl_timer.pack(pady=(0, 10))

    def on_avanzar():
        motivo["valor"] = "avanzar"
        root.destroy()

    def on_cerrar():
        motivo["valor"] = "cerrada"
        root.destroy()

    btn = tk.Button(
        root, text="Avanzar", width=14, height=2,
        font=("Segoe UI", 10, "bold"), command=on_avanzar,
    )
    btn.pack(pady=10)

    def tick():
        if not root.winfo_exists():
            return
        restante["seg"] -= 1
        if restante["seg"] <= 0:
            try:
                root.destroy()
            except Exception:
                pass
            return
        lbl_timer.config(text=f"Auto-avanza en {restante['seg']} s")
        root.after(1000, tick)

    root.protocol("WM_DELETE_WINDOW", on_cerrar)
    root.after(1000, tick)
    root.after(timeout_seg * 1000 + 50,
              lambda: root.winfo_exists() and root.destroy())

    try:
        btn.focus_set()
    except Exception:
        pass

    root.mainloop()
    return motivo["valor"]


def _ventana_avanzar_consola(timeout_seg: int, titulo_caso: str) -> str:
    """
    Fallback en consola cuando tkinter no está disponible.
    """
    print("\n" + "─" * 60)
    print(f"  ⏸   Pausa — {titulo_caso[:80]}")
    print(f"        Presiona ENTER para avanzar  (auto-avanza en {timeout_seg}s)")
    print("─" * 60)

    # Variante Windows
    try:
        import msvcrt
        inicio = time.time()
        ultimo_print = -1
        while True:
            restante = timeout_seg - int(time.time() - inicio)
            if restante <= 0:
                print("\r  ⏱   Tiempo agotado — avanzando…              ")
                return "timeout"
            if restante != ultimo_print:
                sys.stdout.write(f"\r  ⏱   Auto-avanza en {restante:2d}s  (ENTER para continuar)  ")
                sys.stdout.flush()
                ultimo_print = restante
            if msvcrt.kbhit():
                ch = msvcrt.getch()
                if ch in (b"\r", b"\n", b" "):
                    print("\r  ▶   ENTER recibido — avanzando…              ")
                    return "avanzar"
            time.sleep(0.1)
    except ImportError:
        pass

    # Variante Linux/Mac
    try:
        import select
        inicio = time.time()
        ultimo_print = -1
        while True:
            restante = timeout_seg - int(time.time() - inicio)
            if restante <= 0:
                print("\r  ⏱   Tiempo agotado — avanzando…              ")
                return "timeout"
            if restante != ultimo_print:
                sys.stdout.write(f"\r  ⏱   Auto-avanza en {restante:2d}s  (ENTER para continuar)  ")
                sys.stdout.flush()
                ultimo_print = restante
            ready, _, _ = select.select([sys.stdin], [], [], 0.1)
            if ready:
                sys.stdin.readline()
                print("\r  ▶   ENTER recibido — avanzando…              ")
                return "avanzar"
    except Exception:
        pass

    # Último recurso
    print(f"  (Sin detección de teclado disponible — espera {timeout_seg}s)")
    time.sleep(timeout_seg)
    return "timeout"


_TKINTER_AVISADO = {"flag": False}


def ventana_avanzar(timeout_seg: int = 20, titulo_caso: str = "") -> str:
    """
    Pausa con botón "Avanzar". Intenta Tkinter; si no está, fallback a consola.
    """
    if timeout_seg <= 0:
        timeout_seg = 20
    try:
        return _ventana_avanzar_tkinter(timeout_seg, titulo_caso)
    except ImportError:
        if not _TKINTER_AVISADO["flag"]:
            print(
                "\n  ⚠️   Tkinter no está disponible en esta instalación de Python.\n"
                "        Usando ventana de consola como alternativa.\n"
                "        (Para tener la ventana gráfica: reinstalar Python\n"
                "         marcando la opción 'tcl/tk and IDLE')."
            )
            _TKINTER_AVISADO["flag"] = True
        return _ventana_avanzar_consola(timeout_seg, titulo_caso)
    except Exception as exc:
        print(f"  ⚠️   Tkinter falló ({exc}). Usando fallback de consola.")
        return _ventana_avanzar_consola(timeout_seg, titulo_caso)


# ===========================================================================
# EXCEL
# ===========================================================================

COL_CAMINO_HEADER = "Camino a la Pantalla"
COL_RESULTADO_HEADER = "Resultado"
COL_DATO_PREFIX = "Dato"

FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def localizar_columnas(ws) -> dict[str, int]:
    cols: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        cols[str(cell.value).strip()] = cell.column

    faltan = []
    if COL_CAMINO_HEADER not in cols:
        faltan.append(COL_CAMINO_HEADER)
    if COL_RESULTADO_HEADER not in cols:
        faltan.append(COL_RESULTADO_HEADER)
    if faltan:
        raise ValueError(
            f"El Excel no contiene la(s) columna(s): {faltan}. "
            f"Headers detectados: {list(cols.keys())}"
        )
    return cols


def leer_datos_fila(ws, fila: int, columnas: dict[str, int]) -> list[str]:
    datos = []
    for n in range(1, 21):
        header = f"{COL_DATO_PREFIX}{n}"
        col = columnas.get(header)
        if col is None:
            continue
        valor = ws.cell(row=fila, column=col).value
        if valor is None or str(valor).strip() == "":
            continue
        datos.append(str(valor).strip())
    return datos


def escribir_resultado(ws, fila: int, col_resultado: int, texto: str, ok: bool) -> None:
    cell = ws.cell(row=fila, column=col_resultado, value=texto)
    cell.fill = FILL_OK if ok else FILL_ERROR


# ===========================================================================
# FLUJO POR CASO — SIEMPRE PARTE DESDE EL HAMBURGUESA
# ===========================================================================


def preparar_estado_limpio(page: Page) -> None:
    """
    Asegura que NO haya pantallas, menús o modales abiertos antes de
    intentar abrir el menú hamburguesa para el próximo caso.

    Esto es crítico para que cada caso pueda partir desde cero, incluso
    si el caso anterior terminó en un estado inesperado.
    """
    # a) Modal app-modal
    _cerrar_app_modal_si_existe(page)

    # b) Notificación residual
    try:
        if is_notification_modal_visible(page) is not None:
            print("  🧹  Notificación residual — cerrándola…")
            handle_notification_modal(page, "_residual_")
    except Exception:
        pass

    # c) Pantalla activa abierta
    try:
        page.locator("div.tab.activable.active").first.wait_for(
            state="visible", timeout=1000
        )
        print("  🧹  Pantalla previa aún abierta — cerrándola…")
        cerrar_pantalla_layout(page)
    except PlaywrightTimeoutError:
        pass
    except Exception:
        pass

    # d) Menú abierto
    cerrar_menu_si_abierto(page)

    # e) Pausa breve para que Angular termine de renderizar
    page.wait_for_timeout(ENTRE_CASOS_PAUSA_MS)


def procesar_caso(page: Page, ruta: list[str], datos: list[str]) -> tuple[bool, str]:
    """
    Procesa un caso individual. Retorna (ok, texto_resultado).

    Flujo:
        1. Estado limpio (cierra cualquier pantalla/menú/modal previo).
        2. Verifica que el botón hamburguesa esté visible.
        3. Hace clic en el hamburguesa.
        4. Navega la ruta completa.
        5. Espera la vista.
        6. Captura notificación de error (si la hay) y reporta.
        7. (Futuro) Llena los datos.
        8. Cierra la pantalla.
    """
    ruta_str = " > ".join(ruta)

    # 1. Estado limpio entre casos
    preparar_estado_limpio(page)

    # 2. Verificar que el hamburguesa esté visible
    try:
        page.locator("div.toggle.menu").first.wait_for(state="visible", timeout=5000)
    except PlaywrightTimeoutError:
        return False, "FALLO_NAVEGACION: hamburguesa ≡ no visible (¿quedó algo abierto?)"

    # 3. Abrir hamburguesa (SIEMPRE, sin cortocircuito)
    try:
        abrir_menu_hamburguesa(page)
    except Exception as exc:
        return False, f"FALLO_NAVEGACION: no se pudo abrir el menú ({exc})"

    # 4. Navegar la ruta
    if not navegar_ruta_completa(page, ruta):
        return False, "FALLO_NAVEGACION: no se pudo recorrer la ruta completa"

    # 5. Esperar la vista
    try:
        page.locator("div.tab.activable.active").first.wait_for(
            state="visible", timeout=8000
        )
    except PlaywrightTimeoutError:
        notif = handle_notification_modal(page, ruta_str)
        if notif:
            return False, f"ERROR: {notif}"
        return False, "FALLO_NAVEGACION: la pantalla nunca apareció"

    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", ruta_str)[:MAX_FILENAME_LENGTH]
    screenshot(page, f"caso__{safe}")

    # 6. ¿Hay notificación de error?
    notif = handle_notification_modal(page, ruta_str)
    if notif:
        try:
            cerrar_pantalla_layout(page)
        except Exception:
            pass
        return False, f"ERROR: {notif}"

    # 7. TODO FUTURO: si datos no está vacío → llenar_datos(page, datos)
    if datos:
        print(f"  ℹ️   {len(datos)} dato(s) en la fila — pendiente de implementar el llenado.")

    # 8. Cerrar pantalla
    cerrar_pantalla_layout(page)
    return True, "OK"


# ===========================================================================
# FLUJO PRINCIPAL
# ===========================================================================


def run() -> None:
    print("\n" + "=" * 60)
    print("  Automatización QA — Ejecución de Casos de Prueba")
    print("=" * 60)
    print(f"  URL base       : {config.BASE_URL}")
    print(f"  Headless       : {config.HEADLESS}")
    print(f"  Timeout        : {config.TIMEOUT_MS} ms")
    print(f"  Capturas       : {config.SCREENSHOTS_DIR}")
    print(f"  Excel casos    : {config.EXCEL_CASOS}")
    print(f"  ACTIVA_VENTANA : {config.ACTIVA_VENTANA}")
    print(f"  Tiempo ventana : {config.VENTANA_TIMEOUT_SEG} s")
    print("=" * 60 + "\n")

    if not config.EXCEL_CASOS.exists():
        print(f"❌  No se encontró el Excel: {config.EXCEL_CASOS}")
        sys.exit(1)

    print(f"▶ Abriendo Excel: {config.EXCEL_CASOS.name}")
    wb = load_workbook(config.EXCEL_CASOS)
    ws = wb.active
    columnas = localizar_columnas(ws)
    col_camino = columnas[COL_CAMINO_HEADER]
    col_resultado = columnas[COL_RESULTADO_HEADER]
    total_filas = ws.max_row - 1
    print(f"  ✅  Hoja: '{ws.title}' — {total_filas} caso(s) detectado(s).\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=config.HEADLESS,
            slow_mo=BROWSER_SLOW_MO_MS,
            args=["--start-maximized"],
        )
        context = browser.new_context(viewport=None, ignore_https_errors=True)
        context.set_default_timeout(config.TIMEOUT_MS)
        page = context.new_page()

        exitosos: list[str] = []
        fallidos: list[str] = []

        try:
            print("▶ Cargando URL y login…")
            page.goto(config.BASE_URL, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle", timeout=config.TIMEOUT_MS)
            handle_login(page)
            print()

            for fila in range(2, ws.max_row + 1):
                camino = ws.cell(row=fila, column=col_camino).value
                if camino is None or str(camino).strip() == "":
                    print(f"[Fila {fila}] (vacía) — se omite")
                    continue

                ruta = parse_camino(str(camino))
                ruta_str = " > ".join(ruta)
                print(f"\n{'─'*60}")
                print(f"[Fila {fila}] {ruta_str}")
                print(f"{'─'*60}")

                datos = leer_datos_fila(ws, fila, columnas)

                try:
                    ok, resultado = procesar_caso(page, ruta, datos)
                except Exception as exc:
                    ok, resultado = False, f"EXCEPCION: {exc}"
                    try:
                        screenshot(page, f"EXC_fila{fila}")
                    except Exception:
                        pass
                    # Recuperación agresiva: cerrar todo lo que quede
                    try:
                        preparar_estado_limpio(page)
                    except Exception:
                        pass

                print(f"  → {resultado}")
                escribir_resultado(ws, fila, col_resultado, resultado, ok)

                # Guardar Excel después de cada caso
                try:
                    wb.save(config.EXCEL_CASOS)
                except PermissionError:
                    print(
                        f"  ⚠️   No se pudo guardar el Excel "
                        f"(¿lo tienes abierto en Excel?). Continuando."
                    )

                (exitosos if ok else fallidos).append(ruta_str)

                # Ventana "Avanzar"
                if config.ACTIVA_VENTANA:
                    motivo = ventana_avanzar(
                        timeout_seg=config.VENTANA_TIMEOUT_SEG,
                        titulo_caso=ruta_str,
                    )
                    print(f"  ▶   Continuando ({motivo}).")

            print(f"\n{'='*60}")
            print(f"  ✅  Casos exitosos : {len(exitosos)}")
            print(f"  ❌  Casos fallidos : {len(fallidos)}")
            print(f"{'='*60}")

        except Exception as exc:
            print(f"\n❌  Error fatal: {exc}", file=sys.stderr)
            try:
                screenshot(page, "FATAL_ERROR")
            except Exception:
                pass
            raise
        finally:
            try:
                wb.save(config.EXCEL_CASOS)
                print(f"\n💾  Excel guardado: {config.EXCEL_CASOS}")
            except Exception as exc:
                print(f"\n⚠️   No se pudo guardar el Excel al cerrar: {exc}")
            wb.close()

            context.close()
            browser.close()


if __name__ == "__main__":
    run()
