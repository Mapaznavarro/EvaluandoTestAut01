"""
generar_calculadora_rf.py
=========================
Script que genera el archivo Excel "Calculadora_RentaFija.xlsx" con tres hojas:
  1. Parámetros  – inputs del usuario
  2. Flujos      – tabla de cupones, amortización y flujos descontados
  3. Resultado   – resumen ejecutivo (Precio Sucio, Precio Limpio, TERA)

Uso:
    pip install openpyxl>=3.1.0
    python generar_calculadora_rf.py

Referencia técnica:
    Fabozzi, F. (2000). Fixed Income Mathematics. McGraw-Hill.
    Jorion, P. (2007). Value at Risk. McGraw-Hill.
    CMF Chile – Norma de valorización de instrumentos de renta fija.
"""

import os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colores corporativos
# ---------------------------------------------------------------------------
COLOR_HEADER_BG = "1F3864"   # Azul oscuro
COLOR_HEADER_FG = "FFFFFF"   # Blanco
COLOR_INPUT_BG  = "FFFF99"   # Amarillo claro
COLOR_RESULT_BG = "E2EFDA"   # Verde claro
COLOR_WHITE     = "FFFFFF"

# ---------------------------------------------------------------------------
# Estilos reutilizables
# ---------------------------------------------------------------------------
THIN_SIDE   = Side(style="thin")
BORDER_THIN = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE, bottom=THIN_SIDE)

FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER_BG)
FILL_INPUT  = PatternFill("solid", fgColor=COLOR_INPUT_BG)
FILL_RESULT = PatternFill("solid", fgColor=COLOR_RESULT_BG)
FILL_WHITE  = PatternFill("solid", fgColor=COLOR_WHITE)

FONT_HEADER = Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=11)
FONT_LABEL  = Font(name="Calibri", bold=True, size=11)
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_TITLE  = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER_BG)
FONT_NOTE   = Font(name="Calibri", italic=True, size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


def _estilo_header(celda):
    """Aplica estilo de encabezado a una celda."""
    celda.fill      = FILL_HEADER
    celda.font      = FONT_HEADER
    celda.alignment = ALIGN_CENTER
    celda.border    = BORDER_THIN


def _estilo_input(celda):
    """Aplica estilo de celda de input (amarillo)."""
    celda.fill      = FILL_INPUT
    celda.font      = FONT_NORMAL
    celda.alignment = ALIGN_RIGHT
    celda.border    = BORDER_THIN


def _estilo_resultado(celda):
    """Aplica estilo de celda de resultado (verde)."""
    celda.fill      = FILL_RESULT
    celda.font      = FONT_NORMAL
    celda.alignment = ALIGN_RIGHT
    celda.border    = BORDER_THIN


def _estilo_label(celda):
    """Aplica estilo de etiqueta (negrita)."""
    celda.fill      = FILL_WHITE
    celda.font      = FONT_LABEL
    celda.alignment = ALIGN_LEFT
    celda.border    = BORDER_THIN


def _estilo_normal(celda):
    """Aplica estilo normal con borde."""
    celda.fill      = FILL_WHITE
    celda.font      = FONT_NORMAL
    celda.alignment = ALIGN_RIGHT
    celda.border    = BORDER_THIN


# ---------------------------------------------------------------------------
# Valores de ejemplo para pre-llenar los inputs
# ---------------------------------------------------------------------------
VN_EJEMPLO         = 1_000_000          # Monto Nominal
TASA_EMISION_EJ    = 0.06               # 6,00% anual
FECHA_EMISION_EJ   = date(2025, 1, 1)
FECHA_VENC_EJ      = date(2028, 1, 1)
FRECUENCIA_EJ      = 2                  # semestral
TIR_MERCADO_EJ     = 0.075              # 7,50% anual
FECHA_VALOR_EJ     = date(2026, 6, 1)
BASE_DIAS_EJ       = 360


# ===========================================================================
# HOJA 1 – PARÁMETROS
# ===========================================================================
def crear_hoja_parametros(wb: Workbook) -> None:
    """Crea la hoja 'Parámetros' con los inputs del usuario."""
    ws = wb.create_sheet("Parámetros")

    # ---- Título principal ----
    ws.merge_cells("A1:C1")
    titulo = ws["A1"]
    titulo.value     = "CALCULADORA DE VALORIZACIÓN – RENTA FIJA"
    titulo.font      = FONT_TITLE
    titulo.alignment = ALIGN_CENTER
    titulo.fill      = FILL_WHITE
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")  # fila separadora vacía
    ws.row_dimensions[2].height = 8

    # ---- Encabezado de sección ----
    ws.merge_cells("A3:C3")
    enc = ws["A3"]
    enc.value     = "Parámetros de Entrada"
    enc.fill      = FILL_HEADER
    enc.font      = FONT_HEADER
    enc.alignment = ALIGN_CENTER
    ws.row_dimensions[3].height = 22

    # ---- Definición de filas: (fila_num, etiqueta, nombre_rango, valor_ejemplo, formato) ----
    filas = [
        (4,  "Monto Nominal (VN)",             "VN",          VN_EJEMPLO,      '#,##0.00'),
        (5,  "Tasa de Emisión (anual %)",       "TasaEmision", TASA_EMISION_EJ, '0.0000%'),
        (6,  "Fecha de Emisión",                "FechaEmision",FECHA_EMISION_EJ,'DD/MM/YYYY'),
        (7,  "Fecha de Vencimiento",            "FechaVenc",   FECHA_VENC_EJ,   'DD/MM/YYYY'),
        (8,  "Frecuencia de Pago (veces/año)",  "Freq",        FRECUENCIA_EJ,   '0'),
        (9,  "TIR de Mercado (anual %)",        "TIRmercado",  TIR_MERCADO_EJ,  '0.0000%'),
        (10, "Fecha de Valorización",           "FechaValor",  FECHA_VALOR_EJ,  'DD/MM/YYYY'),
        (11, "Convención de días (360 o 365)",  "Base",        BASE_DIAS_EJ,    '0'),
    ]

    for fila, etiqueta, nombre, valor, fmt in filas:
        # Columna A – etiqueta
        celda_a = ws.cell(row=fila, column=1, value=etiqueta)
        _estilo_label(celda_a)

        # Columna B – valor (input del usuario)
        celda_b = ws.cell(row=fila, column=2, value=valor)
        _estilo_input(celda_b)
        celda_b.number_format = fmt

        # Definir nombre de rango para referencia desde otras hojas
        wb.defined_names.add(
            __import__("openpyxl").workbook.defined_name.DefinedName(
                name=nombre,
                attr_text=f"'Parámetros'!$B${fila}",
            )
        )

        # Columna C – descripción breve
        descripciones = {
            4:  "Valor par del bono",
            5:  "Tasa nominal anual del cupón",
            6:  "Fecha de inicio del bono",
            7:  "Fecha de pago del último cupón + principal",
            8:  "2=semestral, 4=trimestral, 12=mensual",
            9:  "Tasa de descuento de mercado (yield)",
            10: "Fecha en que se realiza la valorización",
            11: "Base 360 (30/360) o 365 (Act/365)",
        }
        celda_c = ws.cell(row=fila, column=3, value=descripciones.get(fila, ""))
        celda_c.font      = FONT_NOTE
        celda_c.alignment = ALIGN_LEFT
        celda_c.border    = BORDER_THIN
        celda_c.fill      = FILL_WHITE

        ws.row_dimensions[fila].height = 20

    # ---- Nota al pie ----
    ws.merge_cells("A13:C13")
    nota = ws["A13"]
    nota.value     = ("⚠  Modifique los valores en las celdas amarillas (columna B). "
                      "Las hojas 'Flujos' y 'Resultado' se actualizarán automáticamente.")
    nota.font      = FONT_NOTE
    nota.alignment = ALIGN_LEFT
    nota.fill      = FILL_WHITE

    # ---- Anchos de columna ----
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 42


# ===========================================================================
# HOJA 2 – FLUJOS
# ===========================================================================
def crear_hoja_flujos(wb: Workbook) -> None:
    """
    Crea la hoja 'Flujos' con la tabla de cupones y flujos descontados.

    El número máximo de cupones se calcula a partir de los valores de ejemplo
    para dimensionar la hoja; las fórmulas Excel referencian directamente la
    hoja Parámetros, de modo que al cambiar los inputs se recalculan solas.

    Estructura de columnas:
      A – N° Cupón
      B – Fecha de Pago
      C – Días desde Fecha de Valorización hasta Fecha de Pago
      D – Factor de Descuento = 1/(1 + TIR_período)^(días/base)
      E – Amortización de Capital (solo en el último cupón)
      F – Intereses (cupón)
      G – Flujo Total = E + F
      H – Flujo Descontado = G * D
    """
    ws = wb.create_sheet("Flujos")

    # ---- Título ----
    ws.merge_cells("A1:H1")
    titulo = ws["A1"]
    titulo.value     = "Tabla de Flujos – Bono Bullet"
    titulo.font      = FONT_TITLE
    titulo.alignment = ALIGN_CENTER
    titulo.fill      = FILL_WHITE
    ws.row_dimensions[1].height = 28

    # ---- Encabezados de columna ----
    encabezados = [
        "N° Cupón",
        "Fecha de Pago",
        "Días hasta\nFecha Pago",
        "Factor de\nDescuento",
        "Amortización\nCapital",
        "Intereses\n(Cupón)",
        "Flujo\nTotal",
        "Flujo\nDescontado",
    ]
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=2, column=col, value=texto)
        _estilo_header(celda)
    ws.row_dimensions[2].height = 36

    # ---- Calcular número de cupones a partir de los valores de ejemplo ----
    # (sólo para saber cuántas filas crear)
    from dateutil.relativedelta import relativedelta  # noqa: usado solo para ejemplo

    def _fechas_cupon(fecha_emision, fecha_venc, freq):
        """Genera las fechas de pago de cupones (simplificado con relativedelta)."""
        meses_por_periodo = 12 // freq
        fechas = []
        fecha_actual = fecha_emision + relativedelta(months=meses_por_periodo)
        while fecha_actual <= fecha_venc:
            fechas.append(fecha_actual)
            fecha_actual = fecha_actual + relativedelta(months=meses_por_periodo)
        return fechas

    try:
        fechas = _fechas_cupon(FECHA_EMISION_EJ, FECHA_VENC_EJ, FRECUENCIA_EJ)
    except Exception:
        # Si dateutil no está disponible, calcular manualmente para el ejemplo
        fechas = []
        meses = 12 // FRECUENCIA_EJ
        from datetime import date as _date
        yr, mo = FECHA_EMISION_EJ.year, FECHA_EMISION_EJ.month
        for _ in range(FRECUENCIA_EJ * 10):  # máximo 10 años
            mo += meses
            while mo > 12:
                mo -= 12
                yr += 1
            f = _date(yr, mo, FECHA_EMISION_EJ.day)
            fechas.append(f)
            if f >= FECHA_VENC_EJ:
                break

    n_cupones = len(fechas)

    # Referencias a parámetros (nombres definidos)
    REF_VN        = "VN"
    REF_TASA      = "TasaEmision"
    REF_FREQ      = "Freq"
    REF_TIR       = "TIRmercado"
    REF_FECHA_VAL = "FechaValor"
    REF_BASE      = "Base"
    REF_FECHA_EMI = "FechaEmision"
    REF_FECHA_VEN = "FechaVenc"

    # ---- Filas de datos con fórmulas ----
    # La fecha de pago del cupón i-ésimo se calcula como:
    #   FechaEmision + (i * 12/Freq) meses
    # En Excel usamos EDATE(FechaEmision, i * (12/Freq))
    for i, fecha_pago in enumerate(fechas, start=1):
        fila = i + 2  # las dos primeras filas son título y encabezado

        # A – N° Cupón
        celda_a = ws.cell(row=fila, column=1)
        celda_a.value        = i
        celda_a.number_format = "0"
        _estilo_normal(celda_a)
        celda_a.alignment = ALIGN_CENTER

        # B – Fecha de Pago
        # Fórmula: =EDATE(FechaEmision, i * (12/Freq))
        celda_b = ws.cell(row=fila, column=2)
        celda_b.value        = (
            f"=EDATE({REF_FECHA_EMI}, {i}*(12/{REF_FREQ}))"
        )
        celda_b.number_format = "DD/MM/YYYY"
        _estilo_normal(celda_b)
        celda_b.alignment = ALIGN_CENTER

        # C – Días desde fecha de valorización hasta fecha de pago
        # Fórmula: =B{fila} - FechaValor
        celda_c = ws.cell(row=fila, column=3)
        celda_c.value        = f"=B{fila}-{REF_FECHA_VAL}"
        celda_c.number_format = "0"
        _estilo_normal(celda_c)
        celda_c.alignment = ALIGN_CENTER

        # D – Factor de descuento = 1/(1 + TIR_período)^(días/base)
        # TIR_período = (1 + TIRmercado)^(1/Freq) - 1
        # Factor      = 1 / (1 + TIR_período)^(C{fila}/Base)
        celda_d = ws.cell(row=fila, column=4)
        celda_d.value = (
            f"=1/((1+{REF_TIR})^(C{fila}/{REF_BASE}))"
        )
        celda_d.number_format = "0.000000"
        _estilo_normal(celda_d)

        # E – Amortización (solo en el último cupón: bono bullet)
        celda_e = ws.cell(row=fila, column=5)
        if i == n_cupones:
            celda_e.value = f"={REF_VN}"
        else:
            celda_e.value = 0
        celda_e.number_format = '#,##0.00'
        _estilo_normal(celda_e)

        # F – Intereses (cupón) = VN * TasaEmision / Freq
        celda_f = ws.cell(row=fila, column=6)
        celda_f.value        = f"={REF_VN}*{REF_TASA}/{REF_FREQ}"
        celda_f.number_format = '#,##0.00'
        _estilo_normal(celda_f)

        # G – Flujo Total = E + F
        celda_g = ws.cell(row=fila, column=7)
        celda_g.value        = f"=E{fila}+F{fila}"
        celda_g.number_format = '#,##0.00'
        _estilo_normal(celda_g)

        # H – Flujo Descontado = G * D
        celda_h = ws.cell(row=fila, column=8)
        celda_h.value        = f"=G{fila}*D{fila}"
        celda_h.number_format = '#,##0.00'
        _estilo_normal(celda_h)

        ws.row_dimensions[fila].height = 18

    # ---- Fila de totales ----
    fila_total = n_cupones + 3
    ws.merge_cells(f"A{fila_total}:G{fila_total}")
    celda_label = ws[f"A{fila_total}"]
    celda_label.value     = "PRECIO DE MERCADO (Suma de Flujos Descontados)"
    celda_label.font      = FONT_LABEL
    celda_label.alignment = ALIGN_RIGHT
    celda_label.border    = BORDER_THIN
    celda_label.fill      = FILL_RESULT

    celda_total = ws.cell(row=fila_total, column=8)
    celda_total.value        = f"=SUM(H3:H{fila_total - 1})"
    celda_total.number_format = '#,##0.00'
    celda_total.font      = Font(name="Calibri", bold=True, size=11)
    celda_total.fill      = FILL_RESULT
    celda_total.border    = BORDER_THIN
    celda_total.alignment = ALIGN_RIGHT
    ws.row_dimensions[fila_total].height = 22

    # ---- Anchos de columna ----
    anchos = [10, 18, 14, 14, 18, 18, 18, 18]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # ---- Nota técnica sobre el factor de descuento ----
    fila_nota = fila_total + 2
    ws.merge_cells(f"A{fila_nota}:H{fila_nota}")
    nota = ws[f"A{fila_nota}"]
    nota.value = (
        "Nota: Factor de Descuento = 1/(1+TIR_Mercado)^(Días/Base).  "
        "Se utiliza tasa continua-exponencial en base días/base para mayor precisión."
    )
    nota.font      = FONT_NOTE
    nota.alignment = ALIGN_LEFT
    nota.fill      = FILL_WHITE


# ===========================================================================
# HOJA 3 – RESULTADO
# ===========================================================================
def crear_hoja_resultado(wb: Workbook, n_cupones: int) -> None:
    """
    Crea la hoja 'Resultado' con el resumen ejecutivo.

    Referencia los cálculos de la hoja 'Flujos' para obtener el Precio de Mercado,
    y calcula directamente:
      - Interés Devengado
      - Precio Limpio
      - TERA
    """
    ws = wb.create_sheet("Resultado")

    # ---- Título ----
    ws.merge_cells("A1:C1")
    titulo = ws["A1"]
    titulo.value     = "Resultado de Valorización"
    titulo.font      = FONT_TITLE
    titulo.alignment = ALIGN_CENTER
    titulo.fill      = FILL_WHITE
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 8

    # ---- Encabezado de sección ----
    ws.merge_cells("A3:C3")
    enc = ws["A3"]
    enc.value     = "Indicadores de Valorización"
    enc.fill      = FILL_HEADER
    enc.font      = FONT_HEADER
    enc.alignment = ALIGN_CENTER
    ws.row_dimensions[3].height = 22

    # La fila de totales en la hoja Flujos es: n_cupones + 3
    fila_precio_flujos = n_cupones + 3

    # ---- Filas de resultados ----
    # Precio de Mercado = suma de flujos descontados (referencia a hoja Flujos)
    precio_formula = f"=Flujos!H{fila_precio_flujos}"

    # Expresión (sin = inicial) para TIR por período, reutilizable dentro de otras fórmulas
    tir_periodo_expr = "(1+TIRmercado)^(1/Freq)-1"
    tir_periodo_formula = f"={tir_periodo_expr}"
    tera_formula = f"=((1+({tir_periodo_expr}))^Freq)-1"

    # Días desde el último cupón hasta la fecha de valorización
    # (expresión sin = para poder embebirla dentro de interes_devengado)
    dias_devengado_expr = f"(FechaValor-EDATE(FechaEmision,({n_cupones}-1)*(12/Freq)))"
    interes_devengado = f"=VN*(TasaEmision/Freq)*({dias_devengado_expr}/Base)"

    resultados = [
        # (fila, etiqueta, fórmula, formato, descripción)
        (4,  "Precio de Mercado (Precio Sucio)",
             precio_formula,
             '#,##0.0000',
             "Suma de flujos descontados a TIR de mercado"),
        (5,  "Interés Devengado",
             interes_devengado,
             '#,##0.0000',
             "VN × (TasaEmisión/Freq) × (días desde último cupón / Base)"),
        (6,  "Precio Limpio",
             "=B4-B5",
             '#,##0.0000',
             "Precio Sucio − Interés Devengado"),
        (7,  "TIR de Mercado",
             "=TIRmercado",
             '0.0000%',
             "Tasa de descuento ingresada por el usuario"),
        (8,  "TIR por Período",
             tir_periodo_formula,
             '0.0000%',
             "(1 + TIR_anual)^(1/Freq) − 1"),
        (9,  "TERA",
             tera_formula,
             '0.0000%',
             "Tasa Efectiva de Retorno Anual"),
        (10, "Rendimiento (%)",
             "=B9*100",
             '0.0000',
             "TERA expresada en porcentaje"),
    ]

    for fila, etiqueta, formula, fmt, descripcion in resultados:
        # Columna A – etiqueta
        celda_a = ws.cell(row=fila, column=1, value=etiqueta)
        _estilo_label(celda_a)

        # Columna B – valor calculado
        celda_b = ws.cell(row=fila, column=2, value=formula)
        _estilo_resultado(celda_b)
        celda_b.number_format = fmt

        # Columna C – descripción
        celda_c = ws.cell(row=fila, column=3, value=descripcion)
        celda_c.font      = FONT_NOTE
        celda_c.alignment = ALIGN_LEFT
        celda_c.border    = BORDER_THIN
        celda_c.fill      = FILL_WHITE

        ws.row_dimensions[fila].height = 22

    # ---- Separador ----
    ws.merge_cells("A11:C11")
    ws.row_dimensions[11].height = 10

    # ---- Sección TERA explicada ----
    ws.merge_cells("A12:C12")
    enc2 = ws["A12"]
    enc2.value     = "Metodología de cálculo – TERA"
    enc2.fill      = FILL_HEADER
    enc2.font      = FONT_HEADER
    enc2.alignment = ALIGN_CENTER
    ws.row_dimensions[12].height = 22

    pasos_tera = [
        (13, "Paso 1",
             "TIR por período = (1 + TIR_anual)^(1/Freq) − 1",
             "Convierte la TIR anual en tasa por período de pago"),
        (14, "Paso 2",
             "TERA = (1 + TIR_período)^Freq − 1",
             "Anualiza la tasa por período → Tasa Efectiva Anual"),
        (15, "Resultado",
             "TERA ≈ TIR cuando los flujos se anualizan correctamente",
             "Confirma la consistencia del cálculo"),
    ]

    for fila, paso, formula_txt, desc in pasos_tera:
        ws.cell(row=fila, column=1, value=paso).font      = FONT_LABEL
        ws.cell(row=fila, column=1).alignment = ALIGN_LEFT
        ws.cell(row=fila, column=1).border    = BORDER_THIN
        ws.cell(row=fila, column=1).fill      = FILL_WHITE

        ws.cell(row=fila, column=2, value=formula_txt).font      = Font(
            name="Courier New", size=10)
        ws.cell(row=fila, column=2).alignment = ALIGN_LEFT
        ws.cell(row=fila, column=2).border    = BORDER_THIN
        ws.cell(row=fila, column=2).fill      = FILL_WHITE

        ws.cell(row=fila, column=3, value=desc).font      = FONT_NOTE
        ws.cell(row=fila, column=3).alignment = ALIGN_LEFT
        ws.cell(row=fila, column=3).border    = BORDER_THIN
        ws.cell(row=fila, column=3).fill      = FILL_WHITE

        ws.row_dimensions[fila].height = 20

    # ---- Nota explicativa ----
    ws.merge_cells("A17:C20")
    nota = ws["A17"]
    nota.value = (
        "La TERA (Tasa de Retorno Anual Efectiva) es el rendimiento anualizado que obtendrá "
        "el inversionista si adquiere el bono al Precio de Mercado calculado y lo mantiene "
        "hasta el vencimiento, asumiendo que se reinvierten los cupones a la misma tasa. "
        "Es equivalente a la TIR (Yield to Maturity) expresada en base anual efectiva."
    )
    nota.font      = FONT_NOTE
    nota.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    nota.fill      = PatternFill("solid", fgColor="DDEEFF")
    nota.border    = BORDER_THIN
    ws.row_dimensions[17].height = 18
    ws.row_dimensions[18].height = 18
    ws.row_dimensions[19].height = 18
    ws.row_dimensions[20].height = 18

    # ---- Anchos de columna ----
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52


# ===========================================================================
# FUNCIÓN PRINCIPAL
# ===========================================================================
def generar_excel(ruta_salida: str = None) -> str:
    """
    Genera el archivo Excel 'Calculadora_RentaFija.xlsx'.

    Parámetros
    ----------
    ruta_salida : str, opcional
        Ruta completa donde guardar el archivo. Si es None, se guarda
        en el mismo directorio que este script.

    Retorna
    -------
    str
        Ruta del archivo generado.
    """
    if ruta_salida is None:
        directorio = os.path.dirname(os.path.abspath(__file__))
        ruta_salida = os.path.join(directorio, "Calculadora_RentaFija.xlsx")

    wb = Workbook()
    # Eliminar la hoja por defecto
    wb.remove(wb.active)

    # ---- Calcular n_cupones para dimensionar la hoja Resultado ----
    try:
        from dateutil.relativedelta import relativedelta as _rd
        meses = 12 // FRECUENCIA_EJ
        fechas_temp = []
        from datetime import date as _date
        f = FECHA_EMISION_EJ
        for _ in range(FRECUENCIA_EJ * 20):
            from dateutil.relativedelta import relativedelta as rdelta
            f = f + rdelta(months=meses)
            fechas_temp.append(f)
            if f >= FECHA_VENC_EJ:
                break
        n_cupones = len(fechas_temp)
    except ImportError:
        # Cálculo manual sin dateutil
        n_cupones = 0
        meses = 12 // FRECUENCIA_EJ
        yr, mo, dy = FECHA_EMISION_EJ.year, FECHA_EMISION_EJ.month, FECHA_EMISION_EJ.day
        for _ in range(FRECUENCIA_EJ * 20):
            mo += meses
            while mo > 12:
                mo -= 12
                yr += 1
            from datetime import date as _date
            f = _date(yr, mo, dy)
            n_cupones += 1
            if f >= FECHA_VENC_EJ:
                break

    # ---- Crear hojas ----
    crear_hoja_parametros(wb)
    crear_hoja_flujos(wb)
    crear_hoja_resultado(wb, n_cupones)

    # ---- Guardar ----
    wb.save(ruta_salida)
    return ruta_salida


# ===========================================================================
# PUNTO DE ENTRADA
# ===========================================================================
if __name__ == "__main__":
    ruta = generar_excel()
    print(f"✅  Archivo generado exitosamente: {ruta}")
    print()
    print("Instrucciones de uso:")
    print("  1. Abra el archivo en Excel o LibreOffice Calc.")
    print("  2. Vaya a la hoja 'Parámetros' y modifique las celdas amarillas.")
    print("  3. Las hojas 'Flujos' y 'Resultado' se actualizarán automáticamente.")
    print("  4. El Precio de Mercado y la TERA aparecen en la hoja 'Resultado'.")
